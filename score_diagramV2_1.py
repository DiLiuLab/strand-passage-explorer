#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
score_diagramV2_1.py  --  Comprehensive diagram explorer for a single link.

Pipeline
--------
1. GENERATE.  Starting from one signed DT code, run N rounds of simplification.
   Each round follows the *same* mechanism as strand_passage_guiV4_0.py:
   ``snappy.Link(dt) -> backtrack_simplify(mode='global') -> export new DT``.
   The randomized backtrack escapes local minima, so each round surfaces a
   (usually different) diagram of the *same* link; that new DT becomes the root
   of the next round.  N rounds => N+1 DT codes (1 initial + N simplified).

2. DEDUP.  Many of those DT codes are the *same diagram* written differently
   (cyclic re-labelling, direction reversal, component reordering, planar
   reflection/flip).  We collapse them to representatives using an exact
   signed-diagram isomorphism test (Weisfeiler-Lehman hash for bucketing +
   VF2 for confirmation) that preserves the over/under (chirality) pattern.

3. SCORE.  Each representative is scored with the score_diagram.py metric
   engine (combinatorial balance, planar-graph symmetry, 2-D Tutte energy,
   3-D sphere energy) and ranked by the composite quality.

4. REPORT.  Write an Excel workbook (one ranked row per representative with all
   metrics, plus a run_info sheet) and an SVG figure (2-D Tutte layout + 3-D
   sphere layout for each representative, captioned with its DT code and score).

Reproducibility: round i is driven by a deterministic per-round seed
``f(base_seed, i)``, so the whole chain is reproducible and the generation step
is resumable from a JSONL checkpoint.

Requires SnapPy (``import snappy``); run under ``sage -python`` on the research
machine for the full toolchain.  Depends on score_diagram.py, link_engine_v4_0.py
and draw_dt_original_labelsV4_5.py sitting next to it.

Usage
-----
    python3 score_diagramV2_1.py                              # defaults: example DT, 99 rounds
    python3 score_diagramV2_1.py --dt "DT: [...]" --rounds 99
    python3 score_diagramV2_1.py --xlsx out.xlsx --svg out.svg --json out.json
    # long runs can be chunked under a shell time limit:
    python3 score_diagramV2_1.py --generate-only --max-seconds 40   # repeat until done
"""

import argparse
import importlib.util
import json
import math
import os
import random
import re
import sys
import time

import numpy as np

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401

import networkx as nx

_HERE = os.path.dirname(os.path.abspath(__file__))


def _find_base(filename):
    # Deliberately does NOT search old_scripts/.  That directory is excluded from
    # git, so a helper resolved there imports fine here and is simply absent in a
    # clone -- the failure is invisible locally and total for everyone else.
    for base in (_HERE, os.getcwd(), os.environ.get("DDOL_DIR", "")):
        if base and os.path.exists(os.path.join(base, filename)):
            return base
    return None


def _load_local(name, filename):
    """Import a sibling module by path, registered in sys.modules (dataclasses need it)."""
    base = _find_base(filename)
    if base is None:
        raise FileNotFoundError("Could not find %s next to score_diagramV2_1.py." % filename)
    if base not in sys.path:
        sys.path.insert(0, base)          # let intra-package `import ...` statements resolve
    spec = importlib.util.spec_from_file_location(name, os.path.join(base, filename))
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    spec.loader.exec_module(mod)
    return mod


def _version_key(path):
    """Sort key for a versioned filename: its trailing V<major>_<minor>... as ints.

    Compares NUMERICALLY.  Plain text ordering would rank a future 'V10_0' below
    'V5_5', because '1' sorts before '5' character-wise.  Accepts every suffix
    spelling in this repo -- 'V5_5' (no separator), '_v4_0', '_V2_0'.  An
    unversioned file yields (), sorting below any versioned one.
    """
    stem = os.path.splitext(os.path.basename(path))[0]
    m = re.search(r"[_-]?[Vv](\d[A-Za-z0-9_]*)$", stem)
    return tuple(int(n) for n in re.findall(r"\d+", m.group(1))) if m else ()


def _find_draw_module():
    """Locate the current draw_dt_original_labels*.py, auto-adapting across version bumps
    (e.g. V4_5 -> V5_5 -> V10_0). Picks the highest-versioned file at the top level."""
    import glob
    for base in (_HERE, os.getcwd(), os.environ.get("DDOL_DIR", "")):
        if not base:
            continue
        matches = glob.glob(os.path.join(base, "draw_dt_original_labels*.py"))
        if matches:
            # basename breaks ties, so an equal-version pair resolves deterministically
            path = max(matches, key=lambda p: (_version_key(p), os.path.basename(p)))
            if base not in sys.path:
                sys.path.insert(0, base)
            return os.path.splitext(os.path.basename(path))[0], path
    raise FileNotFoundError(
        "Could not find draw_dt_original_labels*.py next to score_diagramV2_1.py.")


# Load the drawing helper under its real module name so link_engine's
# `import draw_dt_original_labelsV5_5 as D` reuses the same instance.
_draw_name, _draw_path = _find_draw_module()
_spec = importlib.util.spec_from_file_location(_draw_name, _draw_path)
DDOL = importlib.util.module_from_spec(_spec)
sys.modules[_draw_name] = DDOL
_spec.loader.exec_module(DDOL)
LE = _load_local("link_engine_v4_0", "link_engine_v4_0.py")
CDT2 = _load_local("canonical_dt_V2_0", "canonical_dt_V2_0.py")  # canonical form, combinatorial symmetry, and the rotation-system + eigenvalue point-group engine (sigma/i/Sn)

try:
    from scipy.optimize import linear_sum_assignment
    _HAVE_SCIPY = True
except Exception:  # pragma: no cover
    _HAVE_SCIPY = False


# =========================================================================== #
#  Metric engine (self-contained; the standalone score_diagram.py is archived
#  in old_scripts/ as the record-only original, so V2 does not import it).
# =========================================================================== #
def _cv(values):
    """Coefficient of variation (std / mean); 0 means perfectly uniform."""
    v = np.asarray(values, float)
    m = float(np.mean(v))
    if abs(m) < 1e-12:
        return 0.0
    return float(np.std(v) / m)


def _norm_entropy(counts):
    """Shannon entropy of a distribution, normalized to [0, 1]; 1 = perfectly even."""
    c = np.asarray(counts, float)
    s = float(np.sum(c))
    if s <= 0 or len(c) <= 1:
        return 1.0
    p = c / s
    p = p[p > 0]
    H = -float(np.sum(p * np.log(p)))
    return H / math.log(len(c))


def _assignment_residual(X, Y):
    """Mean matched distance between two equal-size point clouds (optimal bijection
    with SciPy, greedy otherwise)."""
    X = np.asarray(X, float)
    Y = np.asarray(Y, float)
    D = np.linalg.norm(X[:, None, :] - Y[None, :, :], axis=2)
    if _HAVE_SCIPY:
        r, c = linear_sum_assignment(D)
        return float(np.mean(D[r, c]))
    used, tot = set(), 0.0
    for i in range(D.shape[0]):
        for j in np.argsort(D[i]):
            if j not in used:
                used.add(j); tot += D[i, j]; break
    return tot / D.shape[0]


def _rotation_2d(theta):
    c, s = math.cos(theta), math.sin(theta)
    return np.array([[c, -s], [s, c]], float)


def _rotation_about_axis(axis, theta):
    a = np.asarray(axis, float)
    a = a / (np.linalg.norm(a) + 1e-15)
    x, y, z = a
    c, s = math.cos(theta), math.sin(theta)
    C = 1.0 - c
    return np.array([
        [c + x * x * C,     x * y * C - z * s, x * z * C + y * s],
        [y * x * C + z * s, c + y * y * C,     y * z * C - x * s],
        [z * x * C - y * s, z * y * C + x * s, c + z * z * C],
    ], float)


def combinatorial_metrics(model):
    crossings = model["crossings"]
    comp_positions = model["comp_positions"]
    comp_of = model["comp_of"]
    over_at = model["over_at"]
    n = len(crossings)
    C = len(comp_positions)

    visit_lengths = [len(cp) for cp in comp_positions]
    strand_cv = _cv(visit_lengths)
    strand_entropy = _norm_entropy(visit_lengths)
    strand_ratio = (max(visit_lengths) / min(visit_lengths)) if min(visit_lengths) else float("inf")

    L = np.zeros((C, C), int)
    n_self = 0
    for cr in crossings:
        ci = comp_of[cr["odd"]]
        cj = comp_of[cr["even"]]
        if ci == cj:
            n_self += 1
            L[ci, ci] += 1
        else:
            L[ci, cj] += 1
            L[cj, ci] += 1
    n_inter = n - n_self
    link_deg = [int(np.sum(L[i]) - L[i, i]) for i in range(C)]
    link_deg_cv = _cv(link_deg) if C > 1 else 0.0

    signs = [1 if cr["even_signed"] > 0 else -1 for cr in crossings]
    n_pos = sum(1 for s in signs if s > 0)
    n_neg = n - n_pos
    sign_imbalance = abs(n_pos - n_neg) / n

    flips = total = 0
    for cp in comp_positions:
        Lc = len(cp)
        if Lc < 2:
            continue
        for i in range(Lc):
            total += 1
            if over_at[cp[i]] != over_at[cp[(i + 1) % Lc]]:
                flips += 1
    alternating_frac = (flips / total) if total else 1.0

    return {
        "n_crossings": n, "n_components": C,
        "strand_visit_lengths": visit_lengths, "strand_length_cv": strand_cv,
        "strand_balance_entropy": strand_entropy, "strand_length_ratio": strand_ratio,
        "n_self_crossings": n_self, "n_inter_crossings": n_inter,
        "linking_matrix": L.tolist(), "linking_degrees": link_deg,
        "linking_degree_cv": link_deg_cv,
        "n_over_neg_convention_pos": n_pos, "n_neg": n_neg,
        "sign_imbalance": sign_imbalance, "alternating_fraction": alternating_frac,
    }


def _is_seg(node):
    return isinstance(node, tuple) and len(node) == 2 and node[0] == "seg"


def _crossing_of_corner(node):
    if isinstance(node, tuple) and not _is_seg(node) and isinstance(node[0], int):
        return node[0]
    return None


def planar_graph_metrics(model, G):
    ok, emb = nx.check_planarity(G)
    if not ok:
        raise RuntimeError("Diagram graph is not planar (DT may be non-realizable).")
    n = len(model["crossings"])
    faces = DDOL.planar_faces(emb)
    face_degrees = []
    for f in faces:
        segs = [x for x in f if _is_seg(x)]
        if segs:
            face_degrees.append(len(segs))
    face_degrees.sort()
    n_bigons = sum(1 for d in face_degrees if d == 2)
    return {
        "euler_faces": len(face_degrees), "euler_faces_expected": n + 2,
        "face_degrees": face_degrees,
        "face_degree_cv": _cv(face_degrees) if face_degrees else 0.0,
        "n_bigons": n_bigons,
        "automorphism_order": None,   # injected by score_diagram (robust canonical count)
        "_emb": emb,
    }


def _signed_automorphism_order(model, G, cap=20000, time_limit=5.0):
    """Order of the diagram's symmetry group that PRESERVES over/under (the meaningful
    'same sequence' symmetry).  Nodes are labelled O/U/S and further coloured by
    Weisfeiler-Lehman refinement so VF2 only tries colour-preserving maps (automorphisms
    always preserve those colours, so the count is exact); a wall-clock guard in the main
    thread prevents pathological VF2 backtracking from hanging."""
    import threading
    H = nx.Graph(G)
    over_at = model["over_at"]
    labels = {}
    for k, cr in enumerate(model["crossings"]):
        oo, oe = bool(over_at[cr["odd"]]), bool(over_at[cr["even"]])
        for role in ("in_o", "out_o"):
            labels[(k, role)] = "O" if oo else "U"
        for role in ("in_e", "out_e"):
            labels[(k, role)] = "O" if oe else "U"
    for node in H.nodes():
        labels.setdefault(node, "S")
    try:
        wl = nx.weisfeiler_lehman_subgraph_hashes(H, iterations=4)
        colors = {node: "%s|%s" % (labels[node], (h[-1] if h else "")) for node, h in wl.items()}
    except Exception:
        colors = labels
    nx.set_node_attributes(H, colors, "_c")
    nm = nx.algorithms.isomorphism.categorical_node_match("_c", None)
    GM = nx.algorithms.isomorphism.GraphMatcher(H, H, node_match=nm)
    count = {"n": 0}

    def _run():
        for _ in GM.isomorphisms_iter():
            count["n"] += 1
            if count["n"] >= cap:
                break

    timed_out = False
    if threading.current_thread() is threading.main_thread():
        import signal

        def _handler(signum, frame):
            raise TimeoutError()

        old = signal.signal(signal.SIGALRM, _handler)
        signal.setitimer(signal.ITIMER_REAL, time_limit)
        try:
            _run()
        except TimeoutError:
            timed_out = True
        finally:
            signal.setitimer(signal.ITIMER_REAL, 0)
            signal.signal(signal.SIGALRM, old)
    else:
        _run()
    return count["n"], count["n"] >= cap, timed_out


def _crossing_centers_2d(model, P):
    n = len(model["crossings"])
    centers = np.zeros((n, 2), float)
    counts = np.zeros(n, float)
    for node, xy in P.items():
        k = _crossing_of_corner(node)
        if k is not None:
            centers[k] += np.asarray(xy, float)
            counts[k] += 1
    counts[counts == 0] = 1
    return centers / counts[:, None]


def _diagram_arcs(model):
    pos_cross = model["pos_cross"]
    arcs = []
    for cp in model["comp_positions"]:
        Lc = len(cp)
        for i in range(Lc):
            arcs.append((pos_cross[cp[i]], pos_cross[cp[(i + 1) % Lc]]))
    return arcs


def _best_rotational_symmetry_2d(centers, kmax=8, thresh=0.12):
    X = np.asarray(centers, float)
    X = X - X.mean(axis=0)
    scale = float(np.sqrt(np.mean(np.sum(X ** 2, axis=1)))) + 1e-15
    best_k, best_score = 1, 0.0
    for k in range(2, kmax + 1):
        Y = X @ _rotation_2d(2 * math.pi / k).T
        resid = _assignment_residual(X, Y) / scale
        score = max(0.0, 1.0 - resid)
        if resid < thresh and k > best_k:
            best_k = k
        best_score = max(best_score, score) if k == 2 else best_score
    if best_k > 1:
        Y = X @ _rotation_2d(2 * math.pi / best_k).T
        best_score = max(0.0, 1.0 - _assignment_residual(X, Y) / scale)
    return best_k, best_score


def geometric_2d_metrics(model, G):
    P = DDOL.compute_positions(G, "tutte")
    centers = _crossing_centers_2d(model, P)
    span = float(np.sqrt(np.mean(np.sum((centers - centers.mean(0)) ** 2, axis=1)))) + 1e-15
    C = centers / span
    arcs = _diagram_arcs(model)
    lengths = np.array([np.linalg.norm(C[a] - C[b]) for a, b in arcs], float)
    edge_cv = _cv(lengths)
    dirichlet_norm = float(np.mean(lengths ** 2) / (np.mean(lengths) ** 2 + 1e-15))
    from collections import defaultdict
    nbrs = defaultdict(list)
    for a, b in arcs:
        nbrs[a].append(b)
        nbrs[b].append(a)
    devs = []
    for k in range(len(C)):
        others = nbrs.get(k, [])
        if len(others) < 2:
            continue
        angs = sorted(math.atan2(*(C[o] - C[k])[::-1]) for o in others)
        gaps = [(angs[(i + 1) % len(angs)] - angs[i]) % (2 * math.pi) for i in range(len(angs))]
        ideal = 2 * math.pi / len(gaps)
        devs.append(math.degrees(math.sqrt(np.mean([(g - ideal) ** 2 for g in gaps]))))
    angle_dev = float(np.mean(devs)) if devs else 0.0
    sym_k, sym_score = _best_rotational_symmetry_2d(centers)
    return {
        "edge_length_cv": edge_cv, "dirichlet_energy_norm": dirichlet_norm,
        "crossing_angle_rms_dev_deg": angle_dev,
        "sym2d_order": sym_k, "sym2d_score": sym_score,
        "_centers2d": centers, "_arcs": arcs,
    }


def _crossing_centers_3d(model, G):
    dirs = DDOL._kamada_3d_unit_directions(G)
    n = len(model["crossings"])
    centers = np.zeros((n, 3), float)
    counts = np.zeros(n, float)
    for node, d in dirs.items():
        k = _crossing_of_corner(node)
        if k is not None:
            centers[k] += np.asarray(d, float)
            counts[k] += 1
    counts[counts == 0] = 1
    centers = centers / counts[:, None]
    norms = np.linalg.norm(centers, axis=1, keepdims=True)
    norms[norms < 1e-12] = 1.0
    return centers / norms


def _riesz_energy(points, s=1.0):
    P = np.asarray(points, float)
    m = len(P)
    e = 0.0
    for i in range(m):
        for j in range(i + 1, m):
            d = np.linalg.norm(P[i] - P[j])
            if d > 1e-12:
                e += 1.0 / d ** s
    return float(e)


def _best_rotational_symmetry_3d(centers, kmax=6, thresh=0.12):
    X = np.asarray(centers, float)
    X = X - X.mean(axis=0)
    scale = float(np.sqrt(np.mean(np.sum(X ** 2, axis=1)))) + 1e-15
    _, _, Vt = np.linalg.svd(X - X.mean(0), full_matrices=False)
    axes = [Vt[i] for i in range(Vt.shape[0])]
    cen = X.mean(0)
    for i in range(len(X)):
        v = X[i] - cen
        if np.linalg.norm(v) > 1e-9:
            axes.append(v)
    best_k, best_score = 1, 0.0
    for axis in axes:
        for k in range(2, kmax + 1):
            Y = X @ _rotation_about_axis(axis, 2 * math.pi / k).T
            resid = _assignment_residual(X, Y) / scale
            score = max(0.0, 1.0 - resid)
            if resid < thresh and k > best_k:
                best_k, best_score = k, score
            elif k == best_k and score > best_score:
                best_score = score
    if best_k == 1:
        best_score = 0.0
    return best_k, best_score


def sphere_3d_metrics(model, G):
    centers = _crossing_centers_3d(model, G)
    n = len(centers)
    thomson = _riesz_energy(centers, s=1.0)
    ref = _riesz_energy(DDOL._fibonacci_sphere_directions(n), s=1.0)
    spread_quality = float(ref / thomson) if thomson > 0 else 1.0
    pos_cross = model["pos_cross"]
    comp_lengths, turning = [], []
    for cp in model["comp_positions"]:
        ks = [pos_cross[p] for p in cp]
        pts = centers[ks]
        m = len(pts)
        if m < 2:
            comp_lengths.append(0.0)
            continue
        clen = 0.0
        for i in range(m):
            clen += math.acos(float(np.clip(np.dot(pts[i], pts[(i + 1) % m]), -1, 1)))
        comp_lengths.append(clen)
        for i in range(m):
            t1 = pts[i] - pts[(i - 1) % m]
            t2 = pts[(i + 1) % m] - pts[i]
            n1, n2 = np.linalg.norm(t1), np.linalg.norm(t2)
            if n1 > 1e-9 and n2 > 1e-9:
                ang = math.acos(float(np.clip(np.dot(t1, t2) / (n1 * n2), -1, 1)))
                turning.append(ang ** 2)
    strand3d_cv = _cv(comp_lengths)
    bending = float(np.sum(turning))
    sym_k, sym_score = _best_rotational_symmetry_3d(centers)
    return {
        "thomson_energy": thomson, "thomson_reference": ref,
        "sphere_spread_quality": spread_quality, "strand3d_length_cv": strand3d_cv,
        "bending_energy": bending, "sym3d_order": sym_k, "sym3d_score": sym_score,
        "_centers3d": centers,
    }


# Edit WEIGHTS to explore which properties should dominate the composite score.
WEIGHTS = {
    "strand_balance": 1.0,     # equal strand lengths across components
    "diagram_symmetry": 1.0,   # combinatorial (sign-aware) + 3-D geometric symmetry
    "face_regularity": 1.0,    # regular planar faces + few bigons (embedding property)
    "sphere_energy": 1.0,      # evenly spread crossings + low bending in 3-D
}


def _quality_scores(m):
    # The 2-D Tutte/geom2d numbers (edge-length CV, Dirichlet energy, crossing-angle
    # deviation, 2-D positional symmetry) depend on which face draw_dt turns to the
    # OUTSIDE (the 'puncture').  When several faces tie for the largest boundary that
    # choice is not intrinsic to the diagram, so those metrics are NOT scored -- they
    # are kept only as descriptive columns.  Every quality below is computed from
    # puncture-independent data: strand lengths, the sign-aware automorphism group,
    # the planar-embedding face spectrum, and the 3-D sphere layout.
    c, g, q3 = m["combinatorial"], m["graph"], m["sphere3d"]
    strand_balance = 0.5 * c["strand_balance_entropy"] + 0.5 * (1.0 / (1.0 + c["strand_length_cv"]))
    aut = g["automorphism_order"]
    sym_comb = 1.0 - 1.0 / max(1, aut)
    diagram_symmetry = np.mean([sym_comb, q3["sym3d_score"]])
    # 'face regularity' replaces the old puncture-dependent 'geometric strain':
    # face-degree CV and the bigon count are properties of the planar embedding
    # (the combinatorial map), independent of the outer-face choice.
    face_regularity = np.mean([
        1.0 / (1.0 + g["face_degree_cv"]),
        1.0 / (1.0 + g["n_bigons"]),
    ])
    sphere_energy = np.mean([q3["sphere_spread_quality"], 1.0 / (1.0 + q3["strand3d_length_cv"])])
    return {
        "strand_balance": float(strand_balance),
        "diagram_symmetry": float(diagram_symmetry),
        "face_regularity": float(face_regularity),
        "sphere_energy": float(sphere_energy),
    }


def score_diagram(dt_string, negative_even="over"):
    comps = DDOL.parse_dt(dt_string)
    model = DDOL.build_model(comps, negative_even=negative_even)
    G = DDOL.build_gadget_graph(model)
    m = {
        "dt": dt_string,
        "combinatorial": combinatorial_metrics(model),
        "graph": planar_graph_metrics(model, G),
        "geom2d": geometric_2d_metrics(model, G),
        "sphere3d": sphere_3d_metrics(model, G),
    }
    # robust, reproducible symmetry order + group (respecting over/under), from the
    # canonical DT relabellings; replaces the fragile VF2 automorphism enumeration.
    m["graph"]["automorphism_order"] = canonical_symmetry(dt_string)
    m["graph"]["symmetry_group"] = canonical_group(dt_string)
    m["quality"] = _quality_scores(m)
    m["composite"] = float(sum(WEIGHTS[k] * m["quality"][k] for k in WEIGHTS) / sum(WEIGHTS.values()))
    return m


def _strip_private(m):
    out = {}
    for k, v in m.items():
        if isinstance(v, dict):
            out[k] = {kk: vv for kk, vv in v.items() if not kk.startswith("_")}
        else:
            out[k] = v
    return out


# --------------------------------------------------------------------------- #
#  1. Generation  (SnapPy global + backtrack, re-rooted each round)
# --------------------------------------------------------------------------- #
def simplify_once(dt, backtrack_rounds, backtrack_steps, seed):
    """One round: SnapPy global simplify with backtrack, return the new DT string."""
    import snappy
    random.seed(seed)
    try:
        np.random.seed(seed & 0x7FFFFFFF)
    except Exception:
        pass
    L = snappy.Link(dt)
    L = LE.backtrack_simplify(snappy, L, mode="global",
                              rounds=backtrack_rounds, steps=backtrack_steps)
    return LE.dt_to_string(LE.parse_dt_any(L.DT_code()))


def _read_checkpoint(path):
    chain = {}
    if path and os.path.exists(path):
        with open(path) as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                chain[int(rec["round"])] = rec["dt"]
    return [chain[i] for i in sorted(chain)] if chain else []


def _append_checkpoint(path, rnd, dt):
    if not path:
        return
    with open(path, "a") as fh:
        fh.write(json.dumps({"round": rnd, "dt": dt}) + "\n")


def generate_chain(dt0, rounds, backtrack_rounds, backtrack_steps, base_seed,
                   checkpoint=None, max_seconds=None, reset_every=0, verbose=True):
    """Return the list of DT strings [dt0, dt1, ..., dt_rounds]; resumable via checkpoint.

    If ``reset_every > 0`` the chain is re-rooted at ``dt0`` after every
    ``reset_every`` rounds (i.e. rounds reset_every+1, 2*reset_every+1, ...
    start again from the original diagram).  This prevents the walk from getting
    trapped cycling among a few common minimal diagrams and re-seeds exploration
    from the canonical starting point."""
    chain = _read_checkpoint(checkpoint)
    if chain and verbose:
        print("  using EXISTING checkpoint '%s': %d round(s) already present (%d DT codes); "
              "generation resumes from there%s."
              % (checkpoint, len(chain) - 1, len(chain),
                 "" if (len(chain) - 1) < rounds else " (already at/beyond the requested rounds — "
                 "no new generation)"), flush=True)
    if not chain:
        chain = [dt0]
        _append_checkpoint(checkpoint, 0, dt0)
    elif chain[0] != dt0:
        raise ValueError("Checkpoint root DT does not match --dt; use a fresh --checkpoint.")

    t0 = time.time()
    while (len(chain) - 1) < rounds:
        i = len(chain)                       # next round index (1..rounds)
        seed = (int(base_seed) * 1000003 + i) & 0x7FFFFFFF
        root = chain[-1]
        if reset_every and i > 1 and ((i - 1) % int(reset_every) == 0):
            root = dt0                       # periodic re-root at the original diagram
        dt_new = simplify_once(root, backtrack_rounds, backtrack_steps, seed)
        chain.append(dt_new)
        _append_checkpoint(checkpoint, i, dt_new)
        if verbose and (i % 20 == 0 or i == rounds):
            print("  round %4d/%d  %.1fs" % (i, rounds, time.time() - t0), flush=True)
        if max_seconds and (time.time() - t0) >= max_seconds:
            break
    return chain


# --------------------------------------------------------------------------- #
#  2. Dedup  (exact signed-diagram isomorphism)
# --------------------------------------------------------------------------- #
def _iso_graph(dt):
    """Gadget graph with node labels encoding over/under; abstract-iso of this
    graph == same diagram up to rotation / reflection / relabelling / cyclic
    permutation / component reorder, while preserving the over/under pattern."""
    comps = DDOL.parse_dt(dt)
    model = DDOL.build_model(comps)
    G = DDOL.build_gadget_graph(model)
    over_at = model["over_at"]
    labels = {}
    for k, cr in enumerate(model["crossings"]):
        over_o = bool(over_at[cr["odd"]])     # is the odd strand over here?
        over_e = bool(over_at[cr["even"]])
        for role in ("in_o", "out_o"):
            labels[(k, role)] = "O" if over_o else "U"
        for role in ("in_e", "out_e"):
            labels[(k, role)] = "O" if over_e else "U"
    for node in G.nodes():
        if node not in labels:
            labels[node] = "S"                # traversal-arc (segment) node
    nx.set_node_attributes(G, labels, "lab")
    return G, model


def canonical_key(dt):
    """Strong, cheap composite signature of a signed diagram:
    (Weisfeiler-Lehman hash of the over/under-labelled diagram graph,
     strand-length spectrum, planar face-degree spectrum).
    Isomorphic diagrams (rotation/reflection/relabel/cyclic-perm/component-reorder,
    over-under preserved) share this signature; collisions between genuinely
    different diagrams are astronomically unlikely, so no per-member VF2 is needed."""
    G, model = _iso_graph(dt)
    wl = nx.weisfeiler_lehman_graph_hash(G, node_attr="lab", iterations=5)
    strand = tuple(sorted(len(cp) for cp in model["comp_positions"]))
    ok, emb = nx.check_planarity(G)
    fdeg = []
    if ok:
        for f in DDOL.planar_faces(emb):
            segs = [x for x in f if _is_seg(x)]
            if segs:
                fdeg.append(len(segs))
    return (wl, strand, tuple(sorted(fdeg)))


def _exact_iso(dtA, dtB):
    GA, _ = _iso_graph(dtA)
    GB, _ = _iso_graph(dtB)
    if GA.number_of_nodes() != GB.number_of_nodes():
        return False
    nm = nx.algorithms.isomorphism.categorical_node_match("lab", "")
    return nx.is_isomorphic(GA, GB, node_match=nm)


# --------------------------------------------------------------------------- #
#  Rigorous DT-native canonical form (authoritative "same diagram" test)
# --------------------------------------------------------------------------- #
# Two DT codes describe the SAME diagram iff, after re-deriving the DT under every
# choice of (component order, per-component base point, per-component traversal
# direction) [and optionally mirror = swapping over/under everywhere], their
# lexicographically-smallest valid DT codes are equal.  This is the exact meaning
# of "redundant due to symmetry / rotation / flipping / cyclic permutation", stated
# natively in DT terms and verifiable by hand.  It is O(product of component lengths)
# per diagram, so it is used for verification, not for bulk dedup.
def _diagram_tours(dt):
    m = DDOL.build_model(DDOL.parse_dt(dt))
    tours = [[(m["pos_cross"][p], bool(m["over_at"][p])) for p in cp]
             for cp in m["comp_positions"]]
    return tours, len(m["crossings"])


def _component_variants(tour, flip):
    base = [(c, (not o) if flip else o) for c, o in tour]
    L = len(base)
    out = []
    for seq in (base, base[::-1]):
        for s in range(L):
            out.append(seq[s:] + seq[:s])
    return out


def _walk_to_dt(walk, n, comp_bounds):
    slots = [[] for _ in range(n)]
    for i, (cr, ov) in enumerate(walk):
        slots[cr].append((i + 1, ov))
    signed = {}
    for lst in slots:
        if len(lst) != 2:
            return None
        (p1, o1), (p2, o2) = lst
        if (p1 & 1) == (p2 & 1):
            return None                        # not a valid DT for this base point
        if p1 & 1:
            oddp, evenp, ev = p1, p2, o2
        else:
            oddp, evenp, ev = p2, p1, o1
        signed[oddp] = (-evenp if ev else evenp)   # convention: negative even = even over
    tup = []
    for lo, hi in comp_bounds:
        odds = sorted(p for p in signed if (p & 1) and lo < p <= hi)
        tup.append(tuple(signed[p] for p in odds))
    return tuple(tup)


def canonical_dt(dt, allow_flip=True, return_symmetry=False):
    """Lexicographically minimal valid signed DT over all relabellings (and, if
    allow_flip, mirror).  Equal canonical_dt  <=>  same diagram.  When return_symmetry
    is True, also return how many valid relabellings reproduce that minimal code, which
    is the diagram's symmetry order (the count of DT re-encodings that coincide with the
    canonical form) - a robust, VF2-free symmetry measure."""
    from itertools import permutations, product
    tours, n = _diagram_tours(dt)
    C = len(tours)
    best = None
    n_min = 0
    for flip in ((False, True) if allow_flip else (False,)):
        var = [_component_variants(tours[ci], flip) for ci in range(C)]
        for perm in permutations(range(C)):
            bounds, off = [], 0
            for ci in perm:
                bounds.append((off, off + len(tours[ci])))
                off += len(tours[ci])
            for choice in product(*[var[ci] for ci in perm]):
                walk = []
                for seq in choice:
                    walk.extend(seq)
                tup = _walk_to_dt(walk, n, bounds)
                if tup is None:
                    continue
                if best is None or tup < best:
                    best, n_min = tup, 1
                elif tup == best:
                    n_min += 1
    return (best, n_min) if return_symmetry else best


def dedup(chain):
    """Collapse identical diagrams.  Returns a list of class dicts sorted by first
    appearance.  Groups by the composite signature from canonical_key(), computed once
    per *distinct* DT string (fast: no per-member graph isomorphism search)."""
    # 1. group member round-indices by exact DT string, preserving first-seen order
    str_members, str_first = {}, {}
    for idx, dt in enumerate(chain):
        if dt not in str_members:
            str_members[dt] = []
            str_first[dt] = idx
        str_members[dt].append(idx)
    unique = sorted(str_members, key=lambda s: str_first[s])

    # 2. signature once per distinct string; group by signature.
    #    The signature (WL hash of the over/under-labelled diagram graph + strand-length
    #    spectrum + face-degree spectrum) is an isomorphism invariant, so identical
    #    diagrams always land together; a collision between genuinely different diagrams
    #    is astronomically unlikely and can be ruled out with --verify (exact VF2) or
    #    canonical_dt().
    groups = {}
    str_strings = {}
    for dt in unique:
        sig = canonical_key(dt)
        groups.setdefault(sig, []).append(dt)

    classes = []
    for sig, strings in groups.items():
        members = sorted(m for s in strings for m in str_members[s])
        classes.append({
            "rep_dt": strings[0],              # earliest first-seen string in the class
            "strings": strings,
            "members": members,
            "multiplicity": len(members),
            "rep_round": min(members),
            "n_distinct_strings": len(strings),
            "sig": sig,
        })
    classes.sort(key=lambda c: c["rep_round"])

    # UP-TO-MIRROR MERGE: collapse classes that are mirror images of each other.
    # The whole V2_0 framework (canonical form + symmetry) works up to mirror, so the
    # diagram COUNT should too.  The two Offset clasps are mirror images, so this takes
    # the set from 5 to 4 (Offset a/b become one entry, an enantiomeric pair).  Keyed by
    # the up-to-mirror canonical DT; the earliest-seen class leads each merged group and
    # carries that canonical as its representative (so scores/names are deterministic).
    by_mirror = {}
    for c in classes:
        mk = _mirror_canonical(c["rep_dt"])
        by_mirror.setdefault(mk, []).append(c)
    merged = []
    for mk, cs in by_mirror.items():
        cs.sort(key=lambda c: c["rep_round"])
        base = cs[0]
        for other in cs[1:]:
            base["strings"] = base["strings"] + other["strings"]
            base["members"] = sorted(set(base["members"]) | set(other["members"]))
        base["multiplicity"] = len(base["members"])
        base["rep_round"] = min(base["members"])
        base["n_distinct_strings"] = len(base["strings"])
        base["mirror_canonical"] = mk
        base["mirror_merged"] = len(cs) > 1
        merged.append(base)
    merged.sort(key=lambda c: c["rep_round"])
    classes = merged
    for j, c in enumerate(classes, start=1):
        c["rep_id"] = j
    return classes


def _mirror_dt(dt):
    """The mirror-image DT code: negate every (even) entry, i.e. switch every crossing's
    over/under.  Used so verification matches the UP-TO-MIRROR de-duplication."""
    comps = CDT2.parse_dt(dt)
    return CDT2.fmt_dt(tuple(tuple(-x for x in c) for c in comps))


def verify_classes(classes, sample=25):
    """Confidence check: within each class, run exact labelled-graph VF2 between the
    representative and up to `sample` other distinct DT strings (0 = all).  Returns a
    per-class report; a False means the fast signature merged non-isomorphic diagrams.

    Because de-duplication now works UP TO MIRROR (a diagram and its mirror image share
    a class -- e.g. the two Offset clasps), a member counts as consistent if it is VF2-
    isomorphic to the representative EITHER directly OR after mirroring (negating the DT).
    Without this, VF2 -- which is over/under-preserving and does NOT see mirror images as
    isomorphic -- falsely reports a MERGE ERROR on any mirror-merged class."""
    import random as _r
    nm = nx.algorithms.isomorphism.categorical_node_match("lab", "")
    report = []
    for c in classes:
        others = [s for s in c["strings"] if s != c["rep_dt"]]
        if sample and len(others) > sample:
            others = _r.Random(0).sample(others, sample)
        Grep, _ = _iso_graph(c["rep_dt"])

        def _iso_upto_mirror(o):
            if nx.is_isomorphic(Grep, _iso_graph(o)[0], node_match=nm):
                return True
            try:                                   # allow the mirror image (up-to-mirror dedup)
                return nx.is_isomorphic(Grep, _iso_graph(_mirror_dt(o))[0], node_match=nm)
            except Exception:  # noqa: BLE001
                return False

        ok = all(_iso_upto_mirror(o) for o in others)
        report.append({"rep_id": c["rep_id"], "checked": len(others), "all_isomorphic": ok})
    return report


def check_sampled(classes, queries):
    """For each query DT code, report whether an equivalent diagram was sampled
    (i.e. its signature matches one of the representative classes)."""
    sigmap = {c["sig"]: c for c in classes}
    out = []
    for q in queries:
        try:
            k = canonical_key(q)
        except Exception as exc:  # noqa: BLE001
            out.append({"dt": q, "sampled": False, "error": str(exc)})
            continue
        c = sigmap.get(k)
        out.append({
            "dt": q,
            "sampled": c is not None,
            "matches_rep_id": (c["rep_id"] if c else None),
            "multiplicity": (c["multiplicity"] if c else 0),
            "first_round": (c["rep_round"] if c else None),
        })
    return out


# --------------------------------------------------------------------------- #
#  3. Score + rank
# --------------------------------------------------------------------------- #
def _jones(dt):
    try:
        import snappy
        return str(snappy.Link(dt).jones_polynomial())
    except Exception:
        return "n/a (needs Sage)"


def _linking_fp(dt):
    """Headless same-link fingerprint: sorted off-diagonal linking numbers."""
    try:
        import snappy
        M = snappy.Link(dt).linking_matrix()
        vals = sorted(int(M[i][j]) for i in range(len(M)) for j in range(len(M)) if i < j)
        return tuple(vals)
    except Exception:
        return None


_CANON_CACHE = None
_CANON_CACHE_PATH = os.environ.get("CANON_CACHE", "canonical_cache.json")


def _canonical_entry(dt, allow_flip=False):
    """Return {'dt': canonical string, 'sym': symmetry order}, memoized on disk so the
    (somewhat expensive) canonicalization is computed once per diagram per run."""
    global _CANON_CACHE
    if _CANON_CACHE is None:
        _CANON_CACHE = {}
        if os.path.exists(_CANON_CACHE_PATH):
            try:
                with open(_CANON_CACHE_PATH) as fh:
                    _CANON_CACHE = json.load(fh)
            except Exception:
                _CANON_CACHE = {}
    key = "N:" + dt
    entry = _CANON_CACHE.get(key)
    if isinstance(entry, dict) and all(k in entry for k in ("dt", "sym", "group")):
        return entry
    # canonical form + symmetry order + symmetry GROUP, via the shared canonical_dt module
    res = CDT2.analyze(dt)
    entry = {"dt": res["canonical"], "sym": int(res["symmetry_order"]),
             "group": _short_group(res.get("element_orders") or [1] * res["symmetry_order"])}
    _CANON_CACHE[key] = entry
    _CANON_CACHE["N:" + entry["dt"]] = entry            # canonical maps to itself
    try:
        with open(_CANON_CACHE_PATH, "w") as fh:
            json.dump(_CANON_CACHE, fh)
    except Exception:
        pass
    return entry


def _short_group(orders):
    """Compact symmetry-group name from the multiset of element orders."""
    n = len(orders)
    if n <= 1:
        return "C1"
    if max(orders) == n:
        return "C%d" % n                                # cyclic (has an order-n element)
    if n == 4 and sorted(orders) == [1, 2, 2, 2]:
        return "C2xC2"                                  # Klein four-group
    if n == 6 and sorted(orders) == [1, 2, 2, 2, 3, 3]:
        return "D3"                                     # dihedral of order 6
    return "order-%d" % n


def canonical_dt_string(dt, allow_flip=False):
    if allow_flip:
        return _mirror_canonical(dt)
    return _canonical_entry(dt)["dt"]


_MIRROR_CANON_CACHE = {}


def _mirror_canonical(dt):
    """UP-TO-MIRROR canonical DT string (V2_0, allow_flip=True): a diagram and its
    mirror image collapse to one code.  Used to merge mirror-image classes (the two
    Offset clasps are mirror images, so up to mirror there are 4 diagrams, not 5)."""
    key = "".join(str(dt).split())
    if key in _MIRROR_CANON_CACHE:
        return _MIRROR_CANON_CACHE[key]
    try:
        comps = [tuple(c) for c in CDT2.parse_dt(dt)]
        canon, _, _ = CDT2.canonicalize(comps, allow_flip=True)
        s = CDT2.fmt_dt(canon)
    except Exception:  # noqa: BLE001
        s = _canonical_entry(dt)["dt"]
    _MIRROR_CANON_CACHE[key] = s
    return s


def canonical_symmetry(dt, allow_flip=False):
    return _canonical_entry(dt)["sym"]


def canonical_group(dt):
    return _canonical_entry(dt)["group"]


def score_representatives(classes):
    """Score each class on its CANONICAL DT code so the geometric/energy metrics are
    reproducible and independent of which sampled DT string first represented the class
    (the layout of two isomorphic-but-differently-labelled codes can differ, which would
    otherwise make the composite score wobble between runs)."""
    scored = []
    for c in classes:
        cdt = c.get("mirror_canonical") or _mirror_canonical(c["rep_dt"])  # up-to-mirror canonical
        c["canonical_dt"] = cdt
        m = score_diagram(cdt)            # m["dt"] == up-to-mirror canonical form
        m["_class"] = c
        m["jones"] = _jones(cdt)
        m["linking_fp"] = _linking_fp(cdt)
        scored.append(m)
    scored.sort(key=lambda m: m["composite"], reverse=True)
    for rank, m in enumerate(scored, start=1):
        m["rank"] = rank
    return scored


# --------------------------------------------------------------------------- #
#  Descriptive names for the five distinct diagrams of the target link
# --------------------------------------------------------------------------- #
# Structural finding (see the comparative report): four of the five diagrams
# contain a Bing-double "clasp" of two components -- a pair that is geometrically
# clasped yet has linking number 0.  In every diagram two "frame" components carry
# 3 crossings each, and the remaining two components split 8 crossings between
# them; the split fixes the clasp's balance (4-4, 5-3, or 6-2).  The fifth diagram
# spreads the same 8 crossings symmetrically ("orthogonally") with no localized
# clasp, which is why it alone reaches C2xC2 symmetry.  Names are keyed to the
# canonical DT code (whitespace-insensitive) so they are stable across runs and
# independent of the tie-order of the two 5-3 variants.
_DIAGRAM_NAMES = {
    "DT:[(-28,-26,-24,-22),(-20,-8,-4),(-6,12,-2),(-10,-16,-14,-18)]":
        ("Orthogonal rosette", "clasp-free; 8 crossings spread symmetrically (4-4); 3D point group C2v"),
    "DT:[(-28,-22,10,18),(-8,24,-20),(12,6,-26,-2),(4,16,-14)]":
        ("Balanced clasp", "Bing-double clasp, balanced 4-4 split; 3D point group Ci (inversion centre)"),
    # The two 5-3 Offset clasps are mirror images of each other, so up to mirror they
    # are ONE diagram (an enantiomeric pair).  Both codes map to the same name; the
    # merged class is represented by the up-to-mirror canonical (the second code).
    "DT:[(-28,-26,8),(14,22,-2),(-24,-20,-12),(-16,6,4,-18,-10)]":
        ("Offset clasp", "Bing-double clasp, offset 5-3 split (mirror-image pair a/b); 3D point group Cs (mirror plane)"),
    "DT:[(-28,-26,-22,12,-20),(-2,24,-18),(-4,-10,14),(8,6,-16)]":
        ("Offset clasp", "Bing-double clasp, offset 5-3 split (mirror-image pair a/b); 3D point group Cs (mirror plane)"),
    "DT:[(-28,-26,-22,14,16,20),(-4,24,-2),(-6,12),(10,8,-18)]":
        ("Lopsided clasp", "Bing-double clasp, lopsided 6-2 split; 3D point group C1"),
}


def _diagram_name(dt):
    """Return (name, clasp-structure) for a canonical DT, or ('','') if unknown."""
    return _DIAGRAM_NAMES.get("".join(str(dt).split()), ("", ""))


# --------------------------------------------------------------------------- #
#  4a. Excel report
# --------------------------------------------------------------------------- #
def _cols():
    # (header, path-getter) ; getter takes the scored metrics dict m
    def g(*keys):
        def _get(m):
            d = m
            for k in keys:
                d = d[k]
            return d
        return _get
    return [
        ("Rank", lambda m: m["rank"]),
        ("Rep ID", lambda m: m["_class"]["rep_id"]),
        ("Diagram name", lambda m: _diagram_name(m["dt"])[0]),
        ("Clasp structure", lambda m: _diagram_name(m["dt"])[1]),
        ("Canonical DT code", lambda m: m["dt"]),
        ("Multiplicity", lambda m: m["_class"]["multiplicity"]),
        ("First round", lambda m: m["_class"]["rep_round"]),
        ("Crossings", g("combinatorial", "n_crossings")),
        ("Components", g("combinatorial", "n_components")),
        ("Strand lengths", lambda m: str(m["combinatorial"]["strand_visit_lengths"])),
        ("Strand length CV", g("combinatorial", "strand_length_cv")),
        ("Strand balance entropy", g("combinatorial", "strand_balance_entropy")),
        ("Strand length max/min", g("combinatorial", "strand_length_ratio")),
        ("Self crossings", g("combinatorial", "n_self_crossings")),
        ("Inter crossings", g("combinatorial", "n_inter_crossings")),
        ("Linking degree CV", g("combinatorial", "linking_degree_cv")),
        ("Alternating fraction", g("combinatorial", "alternating_fraction")),
        ("Over/under imbalance", g("combinatorial", "sign_imbalance")),
        ("Faces (n+2)", g("graph", "euler_faces")),
        ("Face degree CV", g("graph", "face_degree_cv")),
        ("Bigons", g("graph", "n_bigons")),
        ("Symmetry order", g("graph", "automorphism_order")),
        ("Symmetry group", lambda m: m["graph"].get("symmetry_group", "")),
        ("Edge length CV", g("geom2d", "edge_length_cv")),
        ("Dirichlet energy", g("geom2d", "dirichlet_energy_norm")),
        ("Crossing angle dev (deg)", g("geom2d", "crossing_angle_rms_dev_deg")),
        ("2D symmetry score", g("geom2d", "sym2d_score")),
        ("Thomson energy", g("sphere3d", "thomson_energy")),
        ("Sphere spread quality", g("sphere3d", "sphere_spread_quality")),
        ("3D strand length CV", g("sphere3d", "strand3d_length_cv")),
        ("Bending energy", g("sphere3d", "bending_energy")),
        ("3D dot-pattern symmetry", g("sphere3d", "sym3d_order")),
        ("3D point group", lambda m: _point_group_3d(m["dt"])),
        ("q: strand balance", g("quality", "strand_balance")),
        ("q: diagram symmetry", g("quality", "diagram_symmetry")),
        ("q: face regularity", g("quality", "face_regularity")),
        ("q: sphere energy", g("quality", "sphere_energy")),
        ("COMPOSITE", lambda m: m["composite"]),
        ("Linking numbers (sorted)", lambda m: str(m.get("linking_fp"))),
        ("Jones (Sage only)", lambda m: m["jones"]),
    ]


# Columns kept for reference but NOT part of the composite score: the 2-D
# Tutte/geom2d numbers depend on which face draw_dt turns to the outside (the
# 'puncture'), so they are descriptive only and shown on a grey background.
_DESCRIPTIVE_COLS = {
    "Edge length CV",
    "Dirichlet energy",
    "Crossing angle dev (deg)",
    "2D symmetry score",
}


def write_excel(scored, path, run_info):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    cols = _cols()
    wb = Workbook()
    ws = wb.active
    ws.title = "representatives"

    head_fill = PatternFill("solid", fgColor="1F3864")
    desc_head_fill = PatternFill("solid", fgColor="808080")   # descriptive-only header
    desc_fill = PatternFill("solid", fgColor="EDEDED")         # descriptive-only cell
    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    best_fill = PatternFill("solid", fgColor="C6EFCE")
    cell_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cidx, (header, _) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=cidx, value=header)
        cell.fill = desc_head_fill if header in _DESCRIPTIVE_COLS else head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for ridx, m in enumerate(scored, start=2):
        best = (m["rank"] == 1)
        for cidx, (header, getter) in enumerate(cols, start=1):
            val = getter(m)
            if isinstance(val, float):
                val = round(val, 4)
            cell = ws.cell(row=ridx, column=cidx, value=val)
            cell.font = cell_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            # Descriptive 2-D columns stay grey even on the best row: they are not
            # part of the composite (puncture-dependent, see _DESCRIPTIVE_COLS).
            if header in _DESCRIPTIVE_COLS:
                cell.fill = desc_fill
            elif best:
                cell.fill = best_fill

    # widths
    for cidx, (header, _) in enumerate(cols, start=1):
        letter = get_column_letter(cidx)
        if header == "DT code":
            ws.column_dimensions[letter].width = 42
        elif header in ("Jones (same link check)", "Strand lengths"):
            ws.column_dimensions[letter].width = 22
        else:
            ws.column_dimensions[letter].width = max(11, min(20, len(header) + 2))
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), len(scored) + 1)

    # run info sheet
    ws2 = wb.create_sheet("run_info")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 80
    for r, (k, v) in enumerate(run_info.items(), start=1):
        a = ws2.cell(row=r, column=1, value=k)
        a.font = Font(name="Arial", bold=True, size=10)
        b = ws2.cell(row=r, column=2, value=str(v))
        b.font = Font(name="Arial", size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top")

    # legend sheet: for every metric column, whether higher or lower is better
    ws3 = wb.create_sheet("metric_legend")
    for col, w in (("A", 26), ("B", 20), ("C", 74)):
        ws3.column_dimensions[col].width = w
    green = PatternFill("solid", fgColor="C6EFCE")   # higher = better
    blue = PatternFill("solid", fgColor="DDEBF7")    # lower = better
    grey = PatternFill("solid", fgColor="EDEDED")    # descriptive / fixed
    for c, txt in enumerate(("Metric", "Direction", "Meaning"), start=1):
        cell = ws3.cell(row=1, column=c, value=txt)
        cell.fill = head_fill
        cell.font = head_font
        cell.border = border
    fillmap = {"higher = better": green, "lower = better": blue,
               "descriptive": grey, "fixed": grey}
    for r, (metric, direction, meaning) in enumerate(_METRIC_LEGEND, start=2):
        for c, val in enumerate((metric, direction, meaning), start=1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            if c == 2:
                cell.fill = fillmap.get(direction, grey)

    wb.save(path)


# For each Excel metric column: is larger better, smaller better, or just descriptive?
_METRIC_LEGEND = [
    ("Rank", "lower = better", "1 = best blueprint overall."),
    ("Diagram name", "descriptive", "Structural name of the diagram: 'Orthogonal rosette' (clasp-free, C2xC2) or one of the Bing-double clasps (Balanced 4-4, Offset 5-3 a/b, Lopsided 6-2)."),
    ("Clasp structure", "descriptive", "The Bing-double reading: two frame components carry 3 crossings each; the two clasp components split 8 crossings; the split (4-4 / 5-3 / 6-2) sets the clasp balance. The orthogonal form has no localized clasp."),
    ("Multiplicity", "descriptive", "How often the simplifier produced this diagram; NOT a quality (the best diagram is actually rare)."),
    ("First round", "descriptive", "Round at which the diagram was first seen."),
    ("Crossings", "lower = better", "Diagram complexity; fixed at 14 here (all are minimal)."),
    ("Components", "fixed", "Number of strands; fixed at 4."),
    ("Strand length CV", "lower = better", "Spread of strand lengths; 0 = all strands equal length."),
    ("Strand balance entropy", "higher = better", "Evenness of strand lengths; 1 = perfectly even."),
    ("Strand length max/min", "lower = better", "Longest / shortest strand; 1 = equal."),
    ("Self crossings", "descriptive", "Within-strand crossings; 0 for this link."),
    ("Inter crossings", "descriptive", "Between-strand crossings."),
    ("Linking degree CV", "lower = better", "Evenness of how the linking load is shared; lower = more even."),
    ("Alternating fraction", "descriptive", "How close to an alternating diagram; not scored."),
    ("Over/under imbalance", "descriptive", "Convention-dependent over/under balance; not scored."),
    ("Faces (n+2)", "fixed", "Number of regions; equals crossings + 2 (Euler check) = 16."),
    ("Face degree CV", "lower = better", "Regularity of the regions; lower = more uniform tiling."),
    ("Bigons", "lower = better", "Two-sided regions (clasps); 0 = no local crowding."),
    ("Symmetry order", "higher = better", "Combinatorial symmetry: how many DT re-encodings fix the diagram (RESPECTS over/under); higher = fewer unique sequences."),
    ("Symmetry group", "descriptive", "The symmetry group type respecting over/under (e.g. C1 trivial, C2, C2xC2 = Klein four-group of order 4, Ck cyclic). C2xC2 has order 4 but is NOT a 4-fold rotation."),
    ("Edge length CV", "descriptive", "DESCRIPTIVE ONLY (not scored): uniformity of segment lengths in the relaxed 2-D layout. Depends on which face draw_dt turns to the outside (the 'puncture'), which is not intrinsic when faces tie for largest -- so it is excluded from the composite."),
    ("Dirichlet energy", "descriptive", "DESCRIPTIVE ONLY (not scored): 2-D spring energy. Puncture-dependent (see Edge length CV), so excluded from the composite."),
    ("Crossing angle dev (deg)", "descriptive", "DESCRIPTIVE ONLY (not scored): deviation of crossings from ideal 90-degree X shapes in the 2-D layout. Puncture-dependent, so excluded from the composite."),
    ("2D symmetry score", "descriptive", "DESCRIPTIVE ONLY (not scored): rotational symmetry of the crossing POSITIONS in the 2-D layout; IGNORES over/under AND depends on the puncture, so excluded from the composite. Use 'Symmetry order' (sign-aware) and '3D point group' instead."),
    ("Thomson energy", "lower = better", "Crowding of crossings on the sphere; lower = more evenly spread."),
    ("Sphere spread quality", "higher = better", "Evenness of the spread on the sphere vs ideal; 1 = ideal."),
    ("3D strand length CV", "lower = better", "Evenness of strand lengths measured on the 3-D sphere."),
    ("Bending energy", "lower = better", "How sharply strands turn in 3-D; lower = gentler, relaxed curves."),
    ("3D dot-pattern symmetry", "higher = better", "Largest rotational symmetry of the crossing POSITIONS only in the 3-D layout (a k-fold dot pattern); IGNORES over/under and crossing signs -- NOT a signed-diagram symmetry, so it can overstate the true symmetry. Use 'Symmetry order'/'Symmetry group' for the sign-respecting value."),
    ("3D point group", "descriptive", "True Schoenflies point group of the 3-D spherical embedding (up to mirror): rotation axes (Cn) classified vs improper operations sub-typed by eigenvalues -- mirror plane (Cs), inversion centre (Ci), rotoreflection (Sn) -- with a full-loop reliability gate. E.g. Orthogonal rosette C2v, Balanced clasp Ci (inversion centre; NOT Cs), Offset clasp Cs, Lopsided clasp C1."),
    ("q: strand balance", "higher = better", "0-1 quality for even strands."),
    ("q: diagram symmetry", "higher = better", "0-1 quality for symmetry (sign-aware automorphism + 3-D dot pattern; the puncture-dependent 2-D symmetry is no longer included)."),
    ("q: face regularity", "higher = better", "0-1 quality for a regular planar tiling: low face-degree CV and few bigons. Computed from the planar embedding (independent of the puncture); replaces the former 2-D 'geometric strain'."),
    ("q: sphere energy", "higher = better", "0-1 quality for even, relaxed 3-D layout."),
    ("COMPOSITE", "higher = better", "Overall 0-1 score; higher = better synthesis blueprint."),
    ("Linking numbers (sorted)", "descriptive", "Pairwise linking fingerprint (same for all: the link is fixed)."),
    ("Jones (Sage only)", "descriptive", "Link invariant; identical for all (same link). Populates under Sage."),
]


# --------------------------------------------------------------------------- #
#  4b. SVG figure  (2-D Tutte layout + 3-D sphere layout per representative)
# --------------------------------------------------------------------------- #
_PALETTE = ["#4c72b0", "#dd8452", "#55a868", "#c44e52", "#8172b3",
            "#937860", "#da8bc3", "#8c8c8c", "#ccb974", "#64b5cd"]


def _arc_list(model):
    """Directed traversal arcs (ci, a, b) between crossing indices, one per visit-step."""
    pos_cross = model["pos_cross"]
    arcs = []
    for ci, cp in enumerate(model["comp_positions"]):
        ks = [pos_cross[p] for p in cp]
        L = len(ks)
        for i in range(L):
            arcs.append((ci, ks[i], ks[(i + 1) % L]))
    return arcs


def _bow_offsets(arcs):
    """Offset multiplier for each arc so that arcs sharing the same endpoint pair
    (e.g. the two arcs of a bigon, or arcs of different components between the same
    crossings) fan apart instead of drawing on top of each other.  Single edges get 0."""
    from collections import defaultdict
    groups = defaultdict(list)
    for idx, (ci, a, b) in enumerate(arcs):
        groups[tuple(sorted((a, b)))].append(idx)
    off = {}
    for _, idxs in groups.items():
        k = len(idxs)
        for j, idx in enumerate(idxs):
            off[idx] = (j - (k - 1) / 2.0)     # symmetric around 0
    return off


def _draw_skeleton_2d(ax, C2, model, gap_frac=0.05):
    """2-D component-coloured skeleton; parallel arcs are bowed apart (quadratic Bezier)."""
    arcs = _arc_list(model)
    off = _bow_offsets(arcs)
    span = float(np.max(C2.max(0) - C2.min(0))) or 1.0
    gap = gap_frac * span
    ts = np.linspace(0, 1, 26)[:, None]
    for idx, (ci, a, b) in enumerate(arcs):
        pa, pb = C2[a], C2[b]
        col = _PALETTE[ci % len(_PALETTE)]
        # Use a CANONICAL direction (sorted endpoints) for the perpendicular so that two
        # parallel arcs between the same crossings bow to OPPOSITE sides regardless of the
        # direction each is traversed.  (Using each arc's own direction flips the perp and
        # made oppositely-traversed bigon arcs overlap.)
        lo, hi = (a, b) if a <= b else (b, a)
        dc = C2[hi] - C2[lo]
        L = float(np.hypot(dc[0], dc[1]))
        if L < 1e-9:
            continue
        perp = np.array([-dc[1], dc[0]]) / L
        ctrl = (pa + pb) / 2.0 + perp * (off[idx] * gap * 2.0)   # midpoint bow = off*gap
        pts = (1 - ts) ** 2 * pa + 2 * (1 - ts) * ts * ctrl + ts ** 2 * pb
        ax.plot(pts[:, 0], pts[:, 1], "-", lw=1.5, color=col, alpha=0.9)
    ax.plot(C2[:, 0], C2[:, 1], "o", ms=4.0, color="#222222", zorder=3)
    ax.set_aspect("equal")
    ax.axis("off")


def _perm_order_local(p):
    seen, order = set(), 1
    for start in range(len(p)):
        if start in seen:
            continue
        L, x = 0, start
        while True:
            seen.add(x); x = p[x]; L += 1
            if x == start:
                break
        order = order * L // math.gcd(order, L)
    return order


_SYM3D_CACHE = {}


def _compute_sym3d(dt, C3):
    """Run the V2_0 rotation-system + eigenvalue engine once for `dt` in the frame
    of C3 and cache both the typed elements and the short point-group label
    (Cn / Cs / Ci / Sn ...).  A full-loop reliability gate is applied inside."""
    try:
        comps = tuple(tuple(c) for c in CDT2.parse_dt(dt))          # dt is already canonical here
        C3 = np.asarray(C3, float)
        res = CDT2.symmetry_3d(comps, centers=C3, loops=CDT2._weave_loops(dt, C3))
        out = {"elements": list(res.get("elements", [])),
               "point_group": res.get("point_group", "").split("(")[0].strip()}
    except Exception:  # noqa: BLE001
        out = {"elements": [], "point_group": ""}
    _SYM3D_CACHE[dt] = out
    return out


def _symmetry_axes_3d(dt, C3):
    """True point-group elements (list of (kind, order, vec); kind in
    {'axis','mirror','inversion','improper-axis'}, vec None for inversion), in the
    frame of C3.  Replaces the old det-only method that mislabelled every improper
    op 'mirror' (e.g. the Balanced clasp's inversion centre)."""
    if dt in _SYM3D_CACHE:
        return _SYM3D_CACHE[dt]["elements"]
    return _compute_sym3d(dt, C3)["elements"]


def _point_group_3d(dt):
    """Short Schoenflies label of the diagram's true 3-D point group (from the
    shared cache; populated once per representative during scoring)."""
    return _SYM3D_CACHE.get(dt, {}).get("point_group", "")


def _draw_symmetry_3d(ax3, sym_elems):
    """Overlay the true point-group elements:
      * rotation axis (Cn)     -> thin short-dashed line + Cn label, extended past the sphere;
      * mirror plane (sigma)   -> translucent square reaching the sphere, thin short-dashed outline;
      * inversion centre (i)   -> a small dot at the centre labelled 'i';
      * rotoreflection (Sn)    -> thin short-dashed axis + Sn label (improper axis)."""
    if not sym_elems:
        return
    from mpl_toolkits.mplot3d.art3d import Poly3DCollection
    dash = (0, (2.4, 1.8))                        # short dashes, small gaps
    # High z-orders + clip_on=False so the axis/label/dot/plane sit ABOVE the faint
    # sphere skeleton (which is drawn at zorder 0) and are never clipped by the panel.
    for kind, order, vec in sym_elems:
        if kind in ("axis", "improper-axis"):
            if vec is None:
                continue
            vec = np.asarray(vec, float); vec = vec / (np.linalg.norm(vec) + 1e-12)
            p0, p1 = -1.6 * vec, 1.6 * vec        # extend out of the sphere
            ax3.plot([p0[0], p1[0]], [p0[1], p1[1]], [p0[2], p1[2]],
                     color="#111111", lw=1.0, ls=dash, alpha=0.95, zorder=30,
                     clip_on=False)
            lbl = ("C%d" % order) if kind == "axis" else ("S%d" % order)
            for s in (1.0, -1.0):                 # label BOTH ends so one is always front-facing
                ax3.text(s * 1.74 * vec[0], s * 1.74 * vec[1], s * 1.74 * vec[2], lbl,
                         color="#111111", fontsize=9, fontweight="bold",
                         ha="center", va="center", zorder=40, clip_on=False)
        elif kind == "inversion":
            # inversion centre: a dot at the origin, labelled i
            ax3.scatter([0.0], [0.0], [0.0], c="#7e3f98", s=46, marker="o",
                        edgecolors="#3a1d47", linewidths=0.6, zorder=32,
                        depthshade=False, clip_on=False)
            ax3.text(0.0, 0.0, 0.18, "i", color="#5b2d70", fontsize=11, fontweight="bold",
                     ha="center", va="center", zorder=40, clip_on=False)
        else:  # mirror plane -> translucent square reaching the sphere, labelled sigma
            if vec is None:
                continue
            n = np.asarray(vec, float); n = n / (np.linalg.norm(n) + 1e-12)
            e1 = np.cross(n, np.array([0.0, 0.0, 1.0]))
            if np.linalg.norm(e1) < 1e-6:
                e1 = np.cross(n, np.array([1.0, 0.0, 0.0]))
            e1 = e1 / np.linalg.norm(e1)
            e2 = np.cross(n, e1); e2 = e2 / (np.linalg.norm(e2) + 1e-12)
            h = 1.05                              # half-side: edges reach the unit sphere
            corners = np.array([h * e1 + h * e2, -h * e1 + h * e2,
                                -h * e1 - h * e2, h * e1 - h * e2])
            poly = Poly3DCollection([corners], facecolor="#7e3f98", edgecolor="none",
                                    alpha=0.06, zorder=1)
            poly.set_clip_on(False)
            ax3.add_collection3d(poly)
            loop = np.vstack([corners, corners[0]])
            ax3.plot(loop[:, 0], loop[:, 1], loop[:, 2],
                     color="#7e3f98", lw=1.0, ls=dash, alpha=0.95, zorder=28,
                     clip_on=False)
            lp = 1.22 * e1                        # sigma label just outside an edge of the square
            ax3.text(lp[0], lp[1], lp[2], "σ", color="#5b2d70", fontsize=10,
                     fontstyle="italic", fontweight="bold", ha="center", va="center",
                     zorder=40, clip_on=False)


def _draw_sphere_depth(ax3, C3, model, elev=22.0, azim=-58.0, gap=0.13, zoom=1.9,
                       sym_elems=None):
    """3-D sphere layout: arcs ride on the sphere (renormalized), parallel arcs are
    bowed apart, and transparency is depth-cued (nearer = opaque, farther = faint)."""
    ax3.view_init(elev=elev, azim=azim)
    er, ar = math.radians(elev), math.radians(azim)
    eye = np.array([math.cos(er) * math.cos(ar),
                    math.cos(er) * math.sin(ar), math.sin(er)], float)
    u = np.linspace(0, 2 * np.pi, 26)
    v = np.linspace(0, np.pi, 13)
    wf = ax3.plot_wireframe(np.outer(np.cos(u), np.sin(v)), np.outer(np.sin(u), np.sin(v)),
                            np.outer(np.ones_like(u), np.cos(v)), color="0.9", linewidth=0.2)
    wf.set_zorder(0); wf.set_clip_on(False)       # skeleton on the lowest layer, never clips
    arcs = _arc_list(model)
    off = _bow_offsets(arcs)
    ss = np.linspace(0, 1, 16)
    for idx, (ci, a, b) in enumerate(arcs):
        pa, pb = C3[a], C3[b]
        col = _PALETTE[ci % len(_PALETTE)]
        # canonical direction (sorted endpoints) so parallel arcs bow to opposite sides
        lo, hi = (a, b) if a <= b else (b, a)
        chord = C3[hi] - C3[lo]
        mid = (pa + pb) / 2.0
        t = np.cross(mid, chord)
        nt = np.linalg.norm(t)
        if nt < 1e-9:
            continue
        t = t / nt
        ctrl = mid + t * (off[idx] * gap * 2.0)
        # bowed arc, each sample renormalized to the unit sphere
        pts = ((1 - ss) ** 2)[:, None] * pa + (2 * (1 - ss) * ss)[:, None] * ctrl \
            + (ss ** 2)[:, None] * pb
        pts = pts / np.linalg.norm(pts, axis=1, keepdims=True)
        for j in range(len(ss) - 1):
            s0, s1 = pts[j], pts[j + 1]
            depth = float(np.dot((s0 + s1) / 2.0, eye))
            alpha = 0.12 + 0.83 * (depth + 1.0) / 2.0
            ax3.plot([s0[0], s1[0]], [s0[1], s1[1]], [s0[2], s1[2]],
                     "-", lw=1.9, color=col, alpha=alpha, solid_capstyle="round",
                     zorder=5, clip_on=False)
    dvals = C3 @ eye
    alphas = 0.2 + 0.8 * (dvals - dvals.min()) / (np.ptp(dvals) + 1e-9)
    for k in range(len(C3)):
        ax3.scatter(C3[k, 0], C3[k, 1], C3[k, 2], c="#222222",
                    s=9, alpha=float(alphas[k]), depthshade=False,
                    zorder=6, clip_on=False)
    _draw_symmetry_3d(ax3, sym_elems)                # overlay C2 axes / mirror planes / i dot
    # EQUAL, symmetric data limits so the sphere renders ROUND.  Without this,
    # Matplotlib auto-scales each axis independently to the plotted data; the symmetry
    # overlays (axis to +-1.6*vec along a diagonal, labels at 1.74, mirror squares to
    # ~1.48) extend the ranges ASYMMETRICALLY, so the cubic box aspect squeezes/stretches
    # the unit sphere.  A fixed cube of half-width 1.8 contains every overlay and keeps
    # all panels the same size.
    _R = 1.8
    ax3.set_xlim(-_R, _R); ax3.set_ylim(-_R, _R); ax3.set_zlim(-_R, _R)
    try:
        ax3.set_box_aspect((1, 1, 1), zoom=zoom)     # zoom fills the panel (bigger sphere)
    except TypeError:
        ax3.set_box_aspect((1, 1, 1))
    ax3.axis("off")


# 2-D layout used for BOTH the draw_dt_original_labels panel and the Tutte skeleton,
# so their orientation/rotation is identical (per request: shaped-tutte, ellipse, aspect 1).
_DRAW_LAYOUT = "shaped-tutte"
_DRAW_TUTTE_OPTS = {"shape": "ellipse", "aspect": 1.0}
_DRAW_MIN_SEP = 0.02          # push apart non-incident strand pieces closer than this (fraction of span)
_SPHERE_VIEWS = [(22.0, -58.0), (22.0, 122.0)]   # two viewpoints (elev, azim), ~180 deg apart


def _render_draw(ax, model, P, centers_d, col_of, show_labels):
    try:
        DDOL.render_diagram(ax, model, P, centers_d, color_of=col_of,
                            show_labels=show_labels, arrows=True, lw=1.7, label_fontsize=5.5)
    except Exception as exc:  # keep the grid robust
        ax.text(0.5, 0.5, "render error:\n%s" % exc, ha="center", va="center",
                fontsize=6, transform=ax.transAxes)
    ax.set_aspect("equal")
    ax.axis("off")


def _strip_clip_paths(fig):
    """Turn off clipping on every artist so the saved SVG contains no <clipPath>
    masks (and nothing is cut off at panel edges).  Iterates all axes and their
    descendant artists, clearing both the clip flag and any assigned clip path."""
    for ax in fig.get_axes():
        try:
            ax.patch.set_clip_on(False)
        except Exception:  # noqa: BLE001
            pass
        for art in ax.findobj():
            try:
                art.set_clip_on(False)
            except Exception:  # noqa: BLE001
                pass
            try:
                art.set_clip_path(None)
            except Exception:  # noqa: BLE001
                pass


def make_figure(scored, path, max_draw=60):
    import textwrap
    reps = scored[:max_draw]
    n = len(reps)
    total = sum(m["_class"]["multiplicity"] for m in scored)
    nsph = len(_SPHERE_VIEWS)
    # columns: text | draw(labelled) | draw(no labels) | skeleton | sphere views...
    ncol = 4 + nsph
    fig = plt.figure(figsize=(3.15 * ncol + 3.6, 3.9 * n + 0.7))
    # wider sphere columns (1.7) so the 3-D drawings are larger; clip is turned off
    # everywhere (see end of this function) so nothing is cut at the panel edges.
    gs = fig.add_gridspec(n, ncol, width_ratios=[1.05, 1.15, 1.15, 0.95] + [1.7] * nsph,
                          hspace=0.30, wspace=0.02)

    for i, m in enumerate(reps):
        model = DDOL.build_model(DDOL.parse_dt(m["dt"]))
        G = DDOL.build_gadget_graph(model)
        # shared shaped-tutte layout -> identical rotation in the drawings and the skeleton
        P = DDOL.compute_positions(G, _DRAW_LAYOUT, tutte_opts=dict(_DRAW_TUTTE_OPTS))
        if _DRAW_MIN_SEP > 0:
            P = DDOL.nudge_min_separation(P, G, _DRAW_MIN_SEP)
        centers_d = DDOL.crossing_centers(model, P)
        C2 = np.array([centers_d[k] for k in range(len(model["crossings"]))], float)
        C3 = m["sphere3d"]["_centers3d"]
        rid = m["_class"]["rep_id"]
        col_of = lambda ci: _PALETTE[ci % len(_PALETTE)]
        group = m["graph"].get("symmetry_group", "C%d" % m["graph"]["automorphism_order"])

        # --- col 0: text (full CANONICAL DT code + metrics) ---
        axt = fig.add_subplot(gs[i, 0])
        axt.axis("off")
        name, clasp = _diagram_name(m["dt"])
        name_line = ("%s\n" % name) if name else ""
        clasp_line = ("%s\n" % "\n".join(textwrap.wrap(clasp, width=34))) if clasp else ""
        head = ("Rank %d   (rep #%d)\n%s%scomposite = %.3f\nmultiplicity = %d / %d\n"
                "symmetry group = %s (order %d)\n3D dot pattern = %d-fold\n"
                "strands = %s\nbending = %.1f"
                % (m["rank"], rid, name_line, clasp_line,
                   m["composite"], m["_class"]["multiplicity"], total,
                   group, m["graph"]["automorphism_order"], m["sphere3d"]["sym3d_order"],
                   m["combinatorial"]["strand_visit_lengths"],
                   m["sphere3d"]["bending_energy"]))
        dt_wrapped = "\n".join(textwrap.wrap(m["dt"], width=30))
        axt.text(0.0, 1.0, head, transform=axt.transAxes, ha="left", va="top",
                 fontsize=9.5, fontweight="bold" if m["rank"] == 1 else "normal")
        axt.text(0.0, 0.30, "Canonical DT code:", transform=axt.transAxes, ha="left", va="top",
                 fontsize=8.5, style="italic")
        axt.text(0.0, 0.23, dt_wrapped, transform=axt.transAxes, ha="left", va="top",
                 fontsize=8.0, family="monospace")

        # --- col 1: draw_dt_original_labels WITH DT labels ---
        axd = fig.add_subplot(gs[i, 1])
        _render_draw(axd, model, P, centers_d, col_of, show_labels=True)
        if i == 0:
            axd.set_title("draw_dt_original_labels\n(with DT labels)", fontsize=9)

        # --- col 2: draw_dt_original_labels WITHOUT labels (strands unobscured) ---
        axd2 = fig.add_subplot(gs[i, 2])
        _render_draw(axd2, model, P, centers_d, col_of, show_labels=False)
        if i == 0:
            axd2.set_title("draw_dt_original_labels\n(no labels)", fontsize=9)

        # --- col 3: 2-D skeleton (same layout, parallel arcs bowed apart) ---
        ax2 = fig.add_subplot(gs[i, 3])
        _draw_skeleton_2d(ax2, C2, model)
        if i == 0:
            ax2.set_title("2-D skeleton\n(same layout)", fontsize=9)

        # --- cols 4+: 3-D sphere, two viewpoints, depth-shaded, enlarged ---
        # symmetry elements (C2 axes / mirror planes) overlaid on the 3-D layout
        sym_elems = _symmetry_axes_3d(m["dt"], C3)
        n_ax = sum(1 for k, o, v in sym_elems if k == "axis")
        n_mir = sum(1 for k, o, v in sym_elems if k == "mirror")
        n_inv = sum(1 for k, o, v in sym_elems if k == "inversion")
        n_S = sum(1 for k, o, v in sym_elems if k == "improper-axis")
        parts = []
        if n_ax:
            parts.append("C%d axis" % max((o for k, o, v in sym_elems if k == "axis"), default=2))
        if n_mir:
            parts.append("%d mirror%s" % (n_mir, "" if n_mir == 1 else "s"))
        if n_inv:
            parts.append("inversion centre (i)")
        if n_S:
            parts.append("%d S%d" % (n_S, max((o for k, o, v in sym_elems if k == "improper-axis"), default=2)))
        sym_label = "3-D symmetry: " + (", ".join(parts) if parts else "none (C1)")
        for vi, (elev, azim) in enumerate(_SPHERE_VIEWS):
            ax3 = fig.add_subplot(gs[i, 4 + vi], projection="3d")
            _draw_sphere_depth(ax3, C3, model, elev=elev, azim=azim, sym_elems=sym_elems)
            if i == 0:
                ax3.set_title("3-D sphere (view %d)\n%s" % (vi + 1, sym_label), fontsize=8)
            elif vi == 0:
                ax3.set_title(sym_label, fontsize=7)

    fig.suptitle("Representative diagrams of one link, ranked by composite score  "
                 "(canonical DT codes at left; 3-D sphere from two viewpoints)",
                 fontsize=13, y=0.999)
    fig.text(0.5, 0.004,
             "On the 3-D spheres (true point group, up to mirror): a black dashed line = a "
             "rotation (Cn) axis, a violet square labelled σ = a mirror plane, a violet dot "
             "labelled i at the centre = an inversion centre, a dashed axis + Sn = a rotoreflection.",
             ha="center", va="bottom", fontsize=8, color="#555555")
    _strip_clip_paths(fig)                   # remove all clipping masks from the SVG
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


def _draw_tutte_opts(puncture=None):
    """Copy of the shared draw options, optionally pinning a specific outer
    ('puncture') face by its crossing-ID signature (a tuple of 'cN' tokens)."""
    opts = dict(_DRAW_TUTTE_OPTS)
    if puncture is not None:
        opts["puncture_face"] = puncture
    return opts


def _draw_one(ax, dt, col_of, show_labels=False, rasterize=False, puncture=None):
    """Render a single DT code with draw_dt_original_labels in its own shaped-tutte
    layout.  ``puncture`` (a crossing-ID signature tuple) pins which face is turned
    to the outside, so the SAME diagram can be drawn with different outer faces."""
    model = DDOL.build_model(DDOL.parse_dt(dt))
    G = DDOL.build_gadget_graph(model)
    P = DDOL.compute_positions(G, _DRAW_LAYOUT, tutte_opts=_draw_tutte_opts(puncture))
    if _DRAW_MIN_SEP > 0:
        P = DDOL.nudge_min_separation(P, G, _DRAW_MIN_SEP)
    centers_d = DDOL.crossing_centers(model, P)
    _render_draw(ax, model, P, centers_d, col_of, show_labels)
    if rasterize:
        ax.set_rasterized(True)      # embed as a small raster in SVG -> fast, valid SVG


def _largest_tie_face_signatures(dt):
    """Crossing-ID signatures of the faces TIED for the largest boundary of a
    diagram's planar embedding -- exactly the faces draw_dt could turn to the
    outside (the 'puncture' choices).  When several tie, each gives a genuinely
    different plane drawing of the SAME diagram; when one face is uniquely
    largest there is a single (canonical) drawing.  Signatures are tuples of
    'cN' tokens, matching draw_dt's puncture-face selector."""
    model = DDOL.build_model(DDOL.parse_dt(dt))
    G = DDOL.build_gadget_graph(model)
    ok, emb = nx.check_planarity(G)
    if not ok:
        return []
    faces = DDOL.planar_faces(emb)
    if not faces:
        return []
    crossing_ids = DDOL.default_crossing_ids(model)
    mx = max(len(f) for f in faces)
    sigs, seen = [], set()
    for f in faces:
        if len(f) != mx:
            continue
        sig = DDOL._face_signature(f, crossing_ids)
        if sig not in seen:
            seen.add(sig)
            sigs.append(sig)
    return sigs


def _puncture_distinct_drawings(dt, want, exclude=None):
    """For one diagram, enumerate the largest-tie puncture faces and return up to
    ``want`` puncture signatures whose drawings are pairwise distinct under
    rotation / flip / strand-reversal, skipping any drawing key in ``exclude``.
    Each returned signature turns a different face to the outside, so the panels
    are the genuinely different plane pictures of the SAME diagram."""
    out, seen = [], set(exclude or ())
    for sig in _largest_tie_face_signatures(dt):
        try:
            key = _draw_congruence_key(dt, puncture=sig)
        except Exception:  # noqa: BLE001
            continue
        if key in seen:
            continue
        seen.add(key)
        out.append(sig)
        if len(out) >= want:
            break
    return out


def _draw_congruence_key(dt, ndigits=3, puncture=None):
    """Signature of a diagram's DRAWING (its draw_dt layout), invariant under planar
    rotation, reflection (flip) and strand-direction reversal -- i.e. exactly the rigid
    moves of the picture, and nothing else.  It is computed ONLY from the rendered
    crossing positions (no DT-canonical form, no graph signature): the sorted, scale-
    normalised multiset of pairwise crossing distances, which is unchanged by rotating,
    mirroring or translating the picture (and by strand reversal, which leaves crossing
    positions put).  Two raw codes share this key iff draw_dt draws them as the same
    picture up to those moves; two codes of the SAME group that draw differently -- e.g.
    with a different face turned to the outside -- get DIFFERENT keys and are both kept."""
    from itertools import combinations
    model = DDOL.build_model(DDOL.parse_dt(dt))
    G = DDOL.build_gadget_graph(model)
    P = DDOL.compute_positions(G, _DRAW_LAYOUT, tutte_opts=_draw_tutte_opts(puncture))
    if _DRAW_MIN_SEP > 0:
        P = DDOL.nudge_min_separation(P, G, _DRAW_MIN_SEP)
    C = DDOL.crossing_centers(model, P)
    pts = np.array([C[k] for k in range(len(model["crossings"]))], float)
    if len(pts) < 2:
        return (len(pts),)
    pts = pts - pts.mean(axis=0)
    d = np.sort(np.array([np.hypot(*(pts[i] - pts[j]))
                          for i, j in combinations(range(len(pts)), 2)]))
    mx = d[-1] if d[-1] > 0 else 1.0
    return tuple(np.round(d / mx, ndigits))


def _distinct_drawings(strings, want, max_scan=80, exclude=None):
    """Pick up to `want` codes whose DRAWINGS are pairwise distinct under
    rotation/flip/strand-reversal, scanning at most `max_scan` codes.  Any drawing
    whose congruence key is in `exclude` is skipped -- used to drop raw codes that
    draw the same as the canonical (which is shown separately)."""
    out, seen = [], set(exclude or ())
    for s in strings[:max_scan]:
        try:
            key = _draw_congruence_key(s)
        except Exception:  # noqa: BLE001
            key = ("err", s)
        if key in seen:
            continue
        seen.add(key)
        out.append(s)
        if len(out) >= want:
            break
    return out


def make_raw_grouping_figure(classes, path, max_per_class=6, max_classes=40, rasterize=False):
    """Show how the many RAW sampled diagrams collapse into the few de-duplicated ones.
    Each row is one class: a text summary, its canonical diagram, then several RAW member
    codes.  The raw members are drawn in their own layout, so they look like rotations /
    reflections of each other -- which is exactly the equivalence used to group them."""
    import textwrap
    reps = classes[:max_classes]
    K = max_per_class
    ncol = 2 + K                                  # text | canonical | puncture drawings...
    nrows = len(reps)
    fig = plt.figure(figsize=(2.7 * ncol + 2.0, 3.25 * nrows + 1.0))
    gs = fig.add_gridspec(nrows, ncol, width_ratios=[1.5, 1.25] + [1.1] * K,
                          hspace=0.62, wspace=0.04)
    col_of = lambda ci: _PALETTE[ci % len(_PALETTE)]

    for i, cl in enumerate(reps):
        canon = canonical_dt_string(cl["rep_dt"])

        # PUNCTURE ENUMERATION: the different plane pictures of a SINGLE diagram come
        # from turning different faces to the outside.  Rather than rely on which raw
        # codes happened to be sampled, enumerate the faces TIED for the largest
        # boundary -- exactly the faces draw_dt could pick as the outer 'puncture' --
        # render the canonical diagram with each, and keep the ones that draw
        # genuinely differently (rotation / flip / strand-reversal).  The canonical is
        # drawn with its default (canonical) puncture in the blue-outlined slot, so we
        # EXCLUDE that drawing; the remaining panels are the other distinct punctures.
        try:
            canon_key = _draw_congruence_key(canon)
        except Exception:  # noqa: BLE001
            canon_key = None
        tie_sigs = _largest_tie_face_signatures(canon)
        sample_sigs = _puncture_distinct_drawings(
            canon, K, exclude={canon_key} if canon_key is not None else None)
        n_shown = 1 + len(sample_sigs)                  # canonical + distinct punctures
        capped = (len(sample_sigs) >= K)                # may be more distinct beyond the cap

        name = _diagram_name(canon)[0]
        head = ("Rank %d — %s" % (i + 1, name)) if name else ("Rank %d" % (i + 1))
        axt = fig.add_subplot(gs[i, 0])
        axt.axis("off")
        txt = ("%s\n(group #%d)\n%d largest-face tie%s\n%d total occurrences\n"
               "%s%d distinct drawing%s shown\n(incl. canonical)\n\ncanonical DT:\n%s"
               % (head, cl["rep_id"], len(tie_sigs), "" if len(tie_sigs) == 1 else "s",
                  cl["multiplicity"],
                  "≥" if capped else "", n_shown, "" if n_shown == 1 else "s",
                  "\n".join(textwrap.wrap(canon.replace("DT: ", ""), 24))))
        axt.text(0.0, 1.0, txt, transform=axt.transAxes, va="top", ha="left", fontsize=8.5)

        axc = fig.add_subplot(gs[i, 1])
        _draw_one(axc, canon, col_of, show_labels=False, rasterize=rasterize)
        for sp in axc.spines.values():
            sp.set_visible(True)
            sp.set_edgecolor("#2c7fb8")
            sp.set_linewidth(2.0)
        axc.set_title("canonical\n(default puncture)\n%s"
                      % "\n".join(textwrap.wrap(canon.replace("DT: ", ""), 24)),
                      fontsize=5.0 if i else 6.0)

        for j in range(K):
            ax = fig.add_subplot(gs[i, 2 + j])
            if j < len(sample_sigs):
                _draw_one(ax, canon, col_of, show_labels=False, rasterize=rasterize,
                          puncture=sample_sigs[j])
                cap = "puncture: " + "+".join(sample_sigs[j])
                ax.set_title("\n".join(textwrap.wrap(cap, 24)), fontsize=5.5)
            else:
                ax.axis("off")

    fig.suptitle("The distinct plane drawings of each diagram, by which face is turned to the outside\n"
                 "(each row = one canonical group; the panels are the genuinely different DRAWINGS "
                 "obtained by puncturing different largest-tie faces, de-duplicated by rotation / flip / "
                 "strand-reversal of the picture)",
                 fontsize=13, y=0.999)
    fig.text(0.5, 0.004,
             "Blue-outlined = canonical representative, drawn with its default puncture.  A diagram "
             "lives on a sphere and draw_dt must turn one face to the outside; when several faces tie "
             "for the largest boundary, each choice gives a different plane picture of the SAME diagram. "
             "Each panel punctures a different tied face (labelled by the crossing IDs on it); drawings "
             "that are rotations / mirrors of one another are merged.  A single tied face (or none shown) "
             "means the diagram has one plane drawing.",
             ha="center", va="bottom", fontsize=8, color="#555555")
    _strip_clip_paths(fig)                   # remove all clipping masks from the SVG
    fig.savefig(path, bbox_inches="tight", dpi=170 if rasterize else 100)
    plt.close(fig)


# --------------------------------------------------------------------------- #
#  5. CLI
# --------------------------------------------------------------------------- #
DEFAULT_DT = "DT: [(-8,-12,16),(-24,-22,-28,-26),(-10,-14,-2),(-20,-6,-18,-4)]"


def launch_gui(defaults=None):
    """Tkinter front-end: fill in the parameters and press Run.  Launched when the
    script is started with no arguments or with --gui.  Falls back to a CLI run if
    Tkinter / a display is unavailable."""
    try:
        import tkinter as tk
        from tkinter import scrolledtext, filedialog, messagebox
        root = tk.Tk()                       # fails here if there is no display
    except Exception as exc:  # no Tk / no display
        print("Tkinter GUI unavailable (%s); running the pipeline on the CLI instead.\n" % exc)
        if defaults is not None:
            defaults.gui = False
            run_pipeline(defaults)
        return

    import threading
    import queue as _queue

    def dv(name, fallback):
        return str(getattr(defaults, name, fallback)) if defaults is not None else str(fallback)

    root.title("DT Diagram Scorer  —  score_diagramV2_1")
    frm = tk.Frame(root, padx=10, pady=8)
    frm.pack(fill="x")

    # GUI default for "Reset to root" is 20 (respects a non-zero value passed via --reset-every).
    reset_default = getattr(defaults, "reset_every", 0) or 20

    # output-file rows get a "Browse..." button (choose folder + file name via a save dialog)
    _save_dialog = {
        "checkpoint": [("JSONL checkpoint", "*.jsonl")],
        "xlsx": [("Excel workbook", "*.xlsx")],
        "svg": [("SVG figure", "*.svg"), ("PNG image", "*.png")],
        "raw_svg": [("SVG figure", "*.svg"), ("PNG image", "*.png")],
        "json": [("JSON", "*.json")],
    }

    def _make_browser(var, filetypes, defext, confirm=True):
        # confirm=False (checkpoint): pick an existing OR new file without the misleading
        # "…already exists, replace?" prompt, since a checkpoint is resumed, not overwritten.
        def _browse():
            path = filedialog.asksaveasfilename(
                title="Choose or open a checkpoint file" if not confirm else "Choose output location",
                defaultextension=defext,
                filetypes=filetypes + [("All files", "*.*")],
                initialfile=os.path.basename(var.get() or ""),
                initialdir=os.path.dirname(var.get() or "") or os.getcwd(),
                confirmoverwrite=confirm)
            if path:
                var.set(path)
        return _browse

    # per-field help shown by the light-blue "?" badge: (title, body-with-example)
    HELP = {
        "dt": ("DT code",
               "The signed Dowker–Thistlethwaite code of the STARTING diagram, grouped by "
               "component. A negative even number marks that the over-strand passes there. Every "
               "alternative the search finds is the SAME link, just drawn differently.\n\n"
               "Example:\nDT: [(-8,-12,16),(-24,-22,-28,-26),(-10,-14,-2),(-20,-6,-18,-4)]"),
        "rounds": ("Rounds (simplifications)",
               "How many simplify-and-perturb cycles to run. Each round yields one alternative "
               "drawing of the same link, so more rounds explore more drawings (and take longer).\n\n"
               "Example: 99 for a quick look, 999 for a thorough search."),
        "backtrack_rounds": ("Backtrack rounds",
               "Inside each round, how many times SnapPy randomly re-tangles then re-simplifies the "
               "diagram to escape the current arrangement. Higher = more variety per round, slower.\n\n"
               "Example: 200."),
        "backtrack_steps": ("Backtrack steps",
               "How many random crossing moves make up one backtrack attempt. Higher = a bigger "
               "perturbation before re-simplifying.\n\nExample: 30."),
        "reset_every": ("Reset to root every N rounds",
               "Re-start the walk from the original DT every N rounds (0 = never). Stops the search "
               "drifting into one region and keeps it sampling the full variety of drawings.\n\n"
               "Example: 20."),
        "seed": ("Seed (reproducibility)",
               "The search is randomized — each round randomly tangles the diagram before "
               "re-simplifying, which is how it finds different diagrams. The seed fixes that "
               "randomness so the run is fully reproducible: the same seed always gives the same "
               "diagrams and scores (and lets a run resume from its checkpoint and be re-run "
               "identically for a paper). Change it to explore a different random path.\n\n"
               "Example: 20260708."),
        "verify": ("Verify sample",
               "An extra EXACT isomorphism check on the grouped diagrams, on top of the fast "
               "signature. 0 = off, -1 = check every member, N = check N per group. Use it to be "
               "certain no genuinely different diagrams were merged. Recomputed every run, even "
               "when resuming from a checkpoint.\n\nExample: 10."),
        "max_draw": ("Max diagrams drawn",
               "Cap on how many ranked representatives appear in the figure. Set it at or above the "
               "number of distinct diagrams the search finds to show them all.\n\nExample: 60."),
        "raw_max_per_class": ("Raw: distinct drawings per group",
               "In the raw-grouping figure, how many DISTINCT drawings to show per group "
               "(de-duplicated by rotation / flip / strand-reversal of the picture; the canonical is "
               "always one of them).\n\nExample: 20."),
        "write_raw": ("Write raw-grouping figure",
               "When ticked, also write the raw-grouping figure — it shows how the many raw sampled "
               "DRAWINGS collapse into the de-duplicated representatives (grouped in rank order, "
               "de-duplicated by rotation / flip / strand-reversal of the picture). Untick it to "
               "skip that figure; its path, per-group count and rasterize option then grey out."),
        "checkpoint": ("Checkpoint file",
               "A JSONL file (use Browse to choose its FOLDER and name) that records EVERY diagram "
               "produced during generation — one line per round. It stores only the generation "
               "chain, not scores or figures.\n\n"
               "• Resume / extend: if the file already exists, generation continues from where it "
               "stopped. Ask for more rounds to add only the difference — e.g. a 999-round file "
               "re-run with Rounds = 1999 generates just the extra 1000 rounds. Asking for the same "
               "or fewer rounds does no new generation.\n"
               "• Always recomputed: dedup, Verify, scoring and ALL figures (including the raw "
               "grouping) are rebuilt fresh from the loaded chain on every run — so changing Verify "
               "or the raw settings and re-running takes effect without re-generating.\n"
               "• Identical continuation needs the same DT, Seed, Backtrack settings and Reset-every "
               "(the root DT must match the file, or the run is refused). A different Seed simply "
               "explores a different additional path.\n\nExample: /path/to/chain999.jsonl."),
        "xlsx": ("Excel output",
               "Path for the metrics workbook (.xlsx); blank = skip. Contains every score, the "
               "diagram names + clasp structure, and a colour-coded metric-direction legend.\n\n"
               "Example: diagram_scores.xlsx."),
        "svg": ("Ranked-figure SVG",
               "Path for the main figure — each ranked diagram drawn four ways (labelled diagram, "
               "2-D skeleton, and 3-D sphere from two views); blank = skip. A .png path also works.\n\n"
               "Example: diagram_scores.svg."),
        "raw_svg": ("Raw grouping SVG",
               "Path for the figure showing how the many raw sampled DRAWINGS collapse into the "
               "de-duplicated representatives (de-duplicated by rotation / flip / strand-reversal of "
               "the picture only); blank = skip. A .png path also works.\n\nExample: grouping.svg."),
        "json": ("JSON output",
               "Path for the full machine-readable results (all metrics + any membership checks); "
               "blank = skip.\n\nExample: results.json."),
        "check": ("Check DT codes (membership test)",
               "After the run, each DT code you paste here (one per line) is tested against the "
               "diagrams that were found: the log reports whether an equivalent diagram was sampled "
               "and which ranked representative it matches (and how often it occurred). Use it to "
               "ask 'is this particular diagram among the ones my search produced?'"),
    }

    def _help_badge(parent, key):
        title, body = HELP[key]
        lbl = tk.Label(parent, text=" ? ", fg="#08306b", bg="#add8e6",
                       font=("TkDefaultFont", 9, "bold"), cursor="hand2",
                       relief="raised", bd=1)
        lbl.bind("<Button-1>", lambda e, t=title, b=body: messagebox.showinfo(t, b))
        return lbl

    vars_ = {}
    widgets = {}          # key -> {"entry": Entry, "browse": Button|None}

    def _full_row(key, label, val, browse=False):
        v = tk.StringVar(value=str(val))
        vars_[key] = v
        row = tk.Frame(frm)
        row.pack(fill="x", pady=2)
        tk.Label(row, text=label, width=26, anchor="w").pack(side="left")
        ent = tk.Entry(row, textvariable=v)
        ent.pack(side="left", fill="x", expand=True, padx=4)
        _help_badge(row, key).pack(side="left", padx=(2, 4))
        btn = None
        if browse and key in _save_dialog:
            ft = _save_dialog[key]
            btn = tk.Button(row, text="Browse…",
                            command=_make_browser(v, ft, ft[0][1].lstrip("*"),
                                                  confirm=(key != "checkpoint")))
            btn.pack(side="left")
        widgets[key] = {"entry": ent, "browse": btn}
        return v

    def _num_pair(spec_a, spec_b):
        # two single-number fields sharing one row (narrow entries)
        row = tk.Frame(frm)
        row.pack(fill="x", pady=2)
        for spec in (spec_a, spec_b):
            if spec is None:
                continue
            key, label, val = spec
            v = tk.StringVar(value=str(val))
            vars_[key] = v
            tk.Label(row, text=label, width=16, anchor="w").pack(side="left")
            ent = tk.Entry(row, textvariable=v, width=11)
            ent.pack(side="left", padx=(0, 2))
            _help_badge(row, key).pack(side="left", padx=(2, 18))
            widgets[key] = {"entry": ent, "browse": None}

    _full_row("dt", "DT code", dv("dt", DEFAULT_DT))
    _num_pair(("rounds", "Rounds", dv("rounds", 99)),
              ("backtrack_rounds", "Backtrack rounds", dv("backtrack_rounds", 200)))
    _num_pair(("backtrack_steps", "Backtrack steps", dv("backtrack_steps", 30)),
              ("reset_every", "Reset every N", reset_default))
    _num_pair(("seed", "Seed", dv("seed", 20260708)),
              ("verify", "Verify 0/-1/N", dv("verify", 10)))
    _num_pair(("max_draw", "Max drawn", dv("max_draw", 60)),
              ("raw_max_per_class", "Raw per group", dv("raw_max_per_class", 20)))
    _full_row("checkpoint", "Checkpoint file", dv("checkpoint", "chainV2.jsonl"), browse=True)
    _full_row("xlsx", "Excel out (blank=skip)",
              getattr(defaults, "xlsx", "") or "diagram_scores.xlsx", browse=True)
    _full_row("svg", "Ranked SVG (blank=skip)",
              getattr(defaults, "svg", "") or "diagram_scores.svg", browse=True)

    # --- raw-grouping figure: a checkbox gates its path + options (dynamic greying) ---
    write_raw_var = tk.BooleanVar(value=bool((getattr(defaults, "raw_svg", "") or "").strip()))
    raw_chk_row = tk.Frame(frm)
    raw_chk_row.pack(fill="x", pady=(6, 0))
    tk.Checkbutton(raw_chk_row, text="Write raw-grouping figure",
                   variable=write_raw_var).pack(side="left")
    _help_badge(raw_chk_row, "write_raw").pack(side="left", padx=6)
    _full_row("raw_svg", "Raw grouping SVG",
              getattr(defaults, "raw_svg", "") or "", browse=True)
    _full_row("json", "JSON out (blank=skip)", getattr(defaults, "json", "") or "", browse=True)

    # --- Check DT codes: membership test (explanation now lives in its "?" badge) ---
    chk_hdr = tk.Frame(frm)
    chk_hdr.pack(fill="x", pady=(10, 0))
    tk.Label(chk_hdr, text="Check DT codes (optional, one DT per line)",
             anchor="w", font=("TkDefaultFont", 10, "bold")).pack(side="left")
    _help_badge(chk_hdr, "check").pack(side="left", padx=6)
    check_text = tk.Text(frm, width=64, height=4)
    check_text.pack(fill="x", pady=2)

    btns = tk.Frame(root, padx=10, pady=4)
    btns.pack(fill="x")
    run_btn = tk.Button(btns, text="Run")
    run_btn.pack(side="left")
    raw_raster_var = tk.BooleanVar(value=bool(getattr(defaults, "raw_raster", False)))
    raw_raster_chk = tk.Checkbutton(
        btns, text="rasterize raw-grouping diagrams (faster/smaller; default = vector)",
        variable=raw_raster_var)
    raw_raster_chk.pack(side="left", padx=14)
    tk.Button(btns, text="Quit", command=root.destroy).pack(side="right")

    # dynamic greying: the raw-grouping path + options are active only when the box is ticked
    def _sync_raw_state(*_):
        state = "normal" if write_raw_var.get() else "disabled"
        for k in ("raw_svg", "raw_max_per_class"):
            w = widgets.get(k, {})
            if w.get("entry") is not None:
                w["entry"].config(state=state)
            if w.get("browse") is not None:
                w["browse"].config(state=state)
        raw_raster_chk.config(state=state)
        if write_raw_var.get() and not vars_["raw_svg"].get().strip():
            vars_["raw_svg"].set("diagram_scores_grouping.svg")   # sensible default when enabled

    write_raw_var.trace_add("write", _sync_raw_state)
    _sync_raw_state()          # apply the initial (default = off) greying

    log = scrolledtext.ScrolledText(root, width=104, height=20, font=("Menlo", 9))
    log.pack(fill="both", expand=True, padx=10, pady=(4, 10))

    q = _queue.Queue()
    _done = object()          # sentinel: worker finished, re-enable the Run button

    class _QWriter:
        def write(self, s):
            q.put(s)

        def flush(self):
            pass

    def _poll():
        # All widget updates happen here on the main thread, driven by the queue.
        try:
            while True:
                item = q.get_nowait()
                if item is _done:
                    run_btn.config(state="normal")     # re-enable for the next run
                else:
                    log.insert("end", item)
                    log.see("end")
        except _queue.Empty:
            pass
        root.after(120, _poll)

    def _run():
        try:
            a = argparse.Namespace(
                dt=vars_["dt"].get().strip() or DEFAULT_DT,
                rounds=int(vars_["rounds"].get()),
                backtrack_rounds=int(vars_["backtrack_rounds"].get()),
                backtrack_steps=int(vars_["backtrack_steps"].get()),
                reset_every=int(vars_["reset_every"].get()),
                seed=int(vars_["seed"].get()),
                checkpoint=vars_["checkpoint"].get().strip() or "chainV2.jsonl",
                max_seconds=0.0, generate_only=False,
                max_draw=int(vars_["max_draw"].get()),
                verify=int(vars_["verify"].get() or 0),
                xlsx=vars_["xlsx"].get().strip() or None,
                svg=vars_["svg"].get().strip() or None,
                raw_svg=(vars_["raw_svg"].get().strip() or None) if write_raw_var.get() else None,
                raw_max_per_class=int(vars_["raw_max_per_class"].get() or 20),
                raw_raster=raw_raster_var.get(),
                json=vars_["json"].get().strip() or None,
                check=[ln.strip() for ln in check_text.get("1.0", "end").splitlines() if ln.strip()],
                check_file=None, gui=False,
            )
        except ValueError as exc:
            q.put("Invalid parameter: %s\n" % exc)
            return
        run_btn.config(state="disabled")
        q.put("\n===== run started =====\n")
        # Import SnapPy now, in the MAIN thread: on first import cypari/cysignals installs
        # signal handlers, which is only allowed in the main thread.  The pipeline runs in a
        # worker thread below, where its own `import snappy` is then a cached no-op.
        try:
            import snappy  # noqa: F401
        except Exception as exc:  # noqa: BLE001
            q.put("(SnapPy not available: %s)\n" % exc)

        def _worker():
            old = sys.stdout
            sys.stdout = _QWriter()
            try:
                run_pipeline(a)
                print("\n===== done =====\n")
            except Exception:  # noqa: BLE001
                import traceback
                print("ERROR:\n" + traceback.format_exc())
            finally:
                sys.stdout = old
                q.put(_done)                           # signal main thread to re-enable

        threading.Thread(target=_worker, daemon=True).start()

    run_btn.config(command=_run)
    q.put("Set parameters and press Run. Long runs stream progress here.\n"
          "Tip: a checkpoint file lets you stop and resume; large 'Rounds' can take minutes.\n")
    root.after(120, _poll)
    root.mainloop()


def main(argv=None):
    raw = list(sys.argv[1:]) if argv is None else list(argv)
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--dt", default=DEFAULT_DT)
    ap.add_argument("--rounds", type=int, default=99)
    ap.add_argument("--backtrack-rounds", type=int, default=200)
    ap.add_argument("--backtrack-steps", type=int, default=30)
    ap.add_argument("--seed", type=int, default=20260708)
    ap.add_argument("--checkpoint", default="chainV2.jsonl")
    ap.add_argument("--reset-every", type=int, default=0,
                    help="re-root the chain at the original DT after every N rounds "
                         "(0 = never); avoids getting trapped cycling among a few diagrams")
    ap.add_argument("--max-seconds", type=float, default=0.0,
                    help="stop generation after this many seconds (0 = no limit); resumable")
    ap.add_argument("--generate-only", action="store_true")
    ap.add_argument("--xlsx", default=None)
    ap.add_argument("--svg", default=None)
    ap.add_argument("--raw-svg", default=None,
                    help="also write an SVG showing how the raw sampled diagrams group into "
                         "the de-duplicated representatives (rotation/flip equivalence)")
    ap.add_argument("--raw-max-per-class", type=int, default=20,
                    help="max distinct drawings shown per group in --raw-svg")
    ap.add_argument("--raw-raster", action="store_true",
                    help="rasterize the diagram panels in --raw-svg (faster/smaller SVG; "
                         "default is full vector art, which is slower to write for many diagrams)")
    ap.add_argument("--json", default=None)
    ap.add_argument("--max-draw", type=int, default=60)
    ap.add_argument("--check", action="append", default=[],
                    help="a DT code to test for membership among the sampled diagrams "
                         "(repeatable)")
    ap.add_argument("--check-file", default=None,
                    help="file with one DT code per line to test for membership")
    ap.add_argument("--verify", type=int, default=10, metavar="N",
                    help="confidence check: exact VF2 of each class rep against up to N "
                         "other distinct members (0 = off; use a large N or -1 for all)")
    ap.add_argument("--gui", action="store_true",
                    help="launch the graphical interface (also the default when no "
                         "arguments are given)")
    args = ap.parse_args(raw)

    if (not raw) or args.gui:
        launch_gui(args)
        return
    run_pipeline(args)


def run_pipeline(args, log=None):
    """Run generation -> dedup -> (verify) -> membership check -> score -> outputs.
    Prints progress with print(); the GUI redirects stdout to capture it."""
    t_start = time.time()
    print("Generating chain: %d rounds, backtrack %dx%d ..."
          % (args.rounds, args.backtrack_rounds, args.backtrack_steps), flush=True)
    chain = generate_chain(
        args.dt, args.rounds, args.backtrack_rounds, args.backtrack_steps,
        args.seed, checkpoint=args.checkpoint,
        max_seconds=(args.max_seconds or None), reset_every=args.reset_every,
    )
    done = len(chain) - 1
    print("Chain has %d/%d rounds (%d DT codes)." % (done, args.rounds, len(chain)), flush=True)
    if args.generate_only or done < args.rounds:
        if done < args.rounds:
            print("Not finished; re-run to resume from checkpoint %s." % args.checkpoint)
        return

    print("Deduplicating %d DT codes ..." % len(chain), flush=True)
    classes = dedup(chain)
    print("  -> %d distinct representative diagrams." % len(classes), flush=True)

    if args.verify:
        n_sample = 0 if args.verify < 0 else args.verify
        rep = verify_classes(classes, sample=n_sample)
        bad = [r for r in rep if not r["all_isomorphic"]]
        print("Verification (exact VF2, sample=%s): %s"
              % (n_sample or "ALL", "all classes internally consistent"
                 if not bad else "MERGE ERROR in classes %s" % [r["rep_id"] for r in bad]),
              flush=True)
        for r in rep:
            print("  class #%d: %d checked, all isomorphic=%s"
                  % (r["rep_id"], r["checked"], r["all_isomorphic"]), flush=True)

    queries = list(args.check)
    if args.check_file and os.path.exists(args.check_file):
        with open(args.check_file) as fh:
            queries += [ln.strip() for ln in fh if ln.strip()]
    check_results = check_sampled(classes, queries) if queries else []
    if check_results:
        print("Membership check (%d queries):" % len(check_results), flush=True)
        for r in check_results:
            if r.get("error"):
                print("  [error] %s : %s" % (r["dt"], r["error"]))
            elif r["sampled"]:
                print("  SAMPLED   -> rep #%d (mult %d, first seen round %d) : %s"
                      % (r["matches_rep_id"], r["multiplicity"], r["first_round"], r["dt"]))
            else:
                print("  NOT found -> %s" % r["dt"])

    print("Scoring representatives ...", flush=True)
    scored = score_representatives(classes)

    # Pre-compute the true 3-D point group once per representative (shared by the
    # xlsx "3D point group" column and the SVG symmetry overlays), in each rep's
    # own 3-D frame so the drawn axes/planes/inversion dot are correctly oriented.
    for m in scored:
        try:
            _compute_sym3d(m["dt"], m["sphere3d"]["_centers3d"])
        except Exception:  # noqa: BLE001
            pass

    fps = set(m["linking_fp"] for m in scored if m["linking_fp"] is not None)
    print("  same-link check: %d distinct linking-number fingerprint(s) among "
          "representatives (expected 1: all are the same link by construction)"
          % len(fps), flush=True)

    # Raw-grouping figure: built AFTER scoring so its groups appear in the SAME order
    # (by composite rank) as the ranked figure, and are labelled with rank + name.
    if getattr(args, "raw_svg", None):
        make_raw_grouping_figure([m["_class"] for m in scored], args.raw_svg,
                                 max_per_class=getattr(args, "raw_max_per_class", 6),
                                 rasterize=getattr(args, "raw_raster", False))
        print("wrote %s" % args.raw_svg, flush=True)

    run_info = {
        "software": "score_diagramV2_1.py",
        "root_DT": args.dt,
        "rounds": args.rounds,
        "backtrack_rounds": args.backtrack_rounds,
        "backtrack_steps": args.backtrack_steps,
        "reset_every": args.reset_every,
        "seed": args.seed,
        "total_DT_codes": len(chain),
        "distinct_representatives": len(classes),
        "distinct_linking_fingerprints": len(fps),
        "jones_polynomial": "populated only under Sage (SnapPy standalone cannot compute it)",
        "scored_on": "canonical DT of each class (labeling-independent, reproducible)",
        "dedup_equivalence": "signed diagram isomorphism (rotation/reflection/relabel/"
                             "cyclic-permutation/component-reorder; over-under preserved). "
                             "Fast signature = WL hash + strand-length + face-degree spectra; "
                             "exact backstops: --verify (VF2) and canonical_dt().",
        "scoring_layout_2d": "Tutte (unit-circle boundary)",
        "figure_layout_2d": "shaped-tutte, ellipse, aspect 1.0 (draw_dt panel and skeleton share it)",
        "layout_3d": "spherical Kamada-Kawai crossing centers (unit sphere)",
        "generated_utc": time.strftime("%Y-%m-%d %H:%M:%S UTC", time.gmtime()),
        "runtime_seconds": round(time.time() - t_start, 1),
    }

    if args.json:
        payload = {
            "run_info": run_info,
            "membership_check": check_results,
            "representatives": [
                {
                    "rep_id": m["_class"]["rep_id"],
                    "rank": m["rank"],
                    "dt": m["dt"],
                    "multiplicity": m["_class"]["multiplicity"],
                    "first_round": m["_class"]["rep_round"],
                    "jones": m["jones"],
                    "linking_fingerprint": list(m["linking_fp"]) if m["linking_fp"] else None,
                    "composite": m["composite"],
                    "quality": m["quality"],
                    "combinatorial": _strip_private({"c": m["combinatorial"]})["c"],
                    "graph": {k: v for k, v in m["graph"].items() if not k.startswith("_")},
                    "geom2d": {k: v for k, v in m["geom2d"].items() if not k.startswith("_")},
                    "sphere3d": {k: v for k, v in m["sphere3d"].items() if not k.startswith("_")},
                }
                for m in scored
            ],
        }
        with open(args.json, "w") as fh:
            json.dump(payload, fh, indent=2)
        print("wrote %s" % args.json)

    if args.xlsx:
        write_excel(scored, args.xlsx, run_info)
        print("wrote %s" % args.xlsx)
    if args.svg:
        make_figure(scored, args.svg, max_draw=args.max_draw)
        print("wrote %s" % args.svg)

    print("\nTop 3 representatives by composite:")
    for m in scored[:3]:
        print("  #%d  composite %.3f  strands %s  |Aut| %d  mult %d"
              % (m["_class"]["rep_id"], m["composite"],
                 m["combinatorial"]["strand_visit_lengths"],
                 m["graph"]["automorphism_order"], m["_class"]["multiplicity"]))


if __name__ == "__main__":
    main()
