#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
score_diagramV2_0.py  --  Comprehensive diagram explorer for a single link.

Pipeline
--------
1. GENERATE.  Starting from one signed DT code, run N rounds of simplification.
   Each round follows the *same* mechanism as strand_passage_guiV4_0.py:
   ``snappy.Link(dt) -> backtrack_simplify(mode='global') -> export new DT``.
   The randomized backtrack escapes local minima, so each round surfaces a
   (usually different) diagram of the *same* link; that new DT becomes the root
   of the next round.  N rounds => N+1 DT codes (1 initial + N simplified).

2. DEDUP.  Many of those DT codes are the *same diagram* written differently
   (cyclic re-labelling, direction reversal, component reordering, planar
   reflection/flip).  We collapse them to representatives using an exact
   signed-diagram isomorphism test (Weisfeiler-Lehman hash for bucketing +
   VF2 for confirmation) that preserves the over/under (chirality) pattern.

3. SCORE.  Each representative is scored with the self-contained V2 metric
   engine (combinatorial balance, planar-graph symmetry, 2-D Tutte energy,
   3-D sphere energy) and ranked by the composite quality.

4. REPORT.  Write an Excel workbook (one ranked row per representative with all
   metrics, plus a run_info sheet) and an SVG figure (2-D Tutte layout + 3-D
   sphere layout for each representative, captioned with its DT code and score).

Reproducibility: round i is driven by a deterministic per-round seed
``f(base_seed, i)``, so the whole chain is reproducible and the generation step
is resumable from a JSONL checkpoint.

Requires SnapPy (``import snappy``); run under ``sage -python`` on the research
machine for the full toolchain.  The metric engine is self-contained and imports
the live ``link_engine_v4_0.py`` and ``draw_dt_original_labelsV5_3.py`` modules
from this folder.

Usage
-----
    sage -python score_diagramV2_0.py                         # open GUI
    sage -python score_diagramV2_0.py --help                  # CLI help
    sage -python score_diagramV2_0.py --dt "DT: [...]" --rounds 99 --xlsx out.xlsx --svg out.svg
    # long runs can be chunked under a shell time limit:
    sage -python score_diagramV2_0.py --generate-only --max-seconds 40   # repeat until done
"""

import argparse
import importlib.util
import json
import math
import os
import random
import sys
import time

import numpy as np

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from mpl_toolkits.mplot3d import Axes3D  # noqa: F401

import networkx as nx

_HERE = os.path.dirname(os.path.abspath(__file__))
SCORE_ICON_PNG = os.path.join(_HERE, "assets", "score_diagram_icon.png")


def _find_base(filename):
    for base in (_HERE, os.getcwd(), os.environ.get("DDOL_DIR", "")):
        if base and os.path.exists(os.path.join(base, filename)):
            return base
    return None


def _load_local(name, filename):
    """Import a sibling module by path, registered in sys.modules (dataclasses need it)."""
    base = _find_base(filename)
    if base is None:
        raise FileNotFoundError("Could not find %s next to score_diagramV2_0.py." % filename)
    if base not in sys.path:
        sys.path.insert(0, base)          # let intra-package `import ...` statements resolve
    spec = importlib.util.spec_from_file_location(name, os.path.join(base, filename))
    mod = importlib.util.module_from_spec(spec)
    sys.modules[name] = mod
    spec.loader.exec_module(mod)
    return mod


DDOL = _load_local("draw_dt_original_labelsV5_3", "draw_dt_original_labelsV5_3.py")
LE = _load_local("link_engine_v4_0", "link_engine_v4_0.py")

try:
    from scipy.optimize import linear_sum_assignment
    _HAVE_SCIPY = True
except Exception:  # pragma: no cover
    _HAVE_SCIPY = False


# =========================================================================== #
#  Metric engine (self-contained; the standalone score_diagram.py is archived
#  in old_scripts/ as the record-only original, so V2 does not import it).
# =========================================================================== #
def _cv(values):
    """Coefficient of variation (std / mean); 0 means perfectly uniform."""
    v = np.asarray(values, float)
    m = float(np.mean(v))
    if abs(m) < 1e-12:
        return 0.0
    return float(np.std(v) / m)


def _norm_entropy(counts):
    """Shannon entropy of a distribution, normalized to [0, 1]; 1 = perfectly even."""
    c = np.asarray(counts, float)
    s = float(np.sum(c))
    if s <= 0 or len(c) <= 1:
        return 1.0
    p = c / s
    p = p[p > 0]
    H = -float(np.sum(p * np.log(p)))
    return H / math.log(len(c))


def _assignment_residual(X, Y):
    """Mean matched distance between two equal-size point clouds (optimal bijection
    with SciPy, greedy otherwise)."""
    X = np.asarray(X, float)
    Y = np.asarray(Y, float)
    D = np.linalg.norm(X[:, None, :] - Y[None, :, :], axis=2)
    if _HAVE_SCIPY:
        r, c = linear_sum_assignment(D)
        return float(np.mean(D[r, c]))
    used, tot = set(), 0.0
    for i in range(D.shape[0]):
        for j in np.argsort(D[i]):
            if j not in used:
                used.add(j); tot += D[i, j]; break
    return tot / D.shape[0]


def _rotation_2d(theta):
    c, s = math.cos(theta), math.sin(theta)
    return np.array([[c, -s], [s, c]], float)


def _rotation_about_axis(axis, theta):
    a = np.asarray(axis, float)
    a = a / (np.linalg.norm(a) + 1e-15)
    x, y, z = a
    c, s = math.cos(theta), math.sin(theta)
    C = 1.0 - c
    return np.array([
        [c + x * x * C,     x * y * C - z * s, x * z * C + y * s],
        [y * x * C + z * s, c + y * y * C,     y * z * C - x * s],
        [z * x * C - y * s, z * y * C + x * s, c + z * z * C],
    ], float)


def combinatorial_metrics(model):
    crossings = model["crossings"]
    comp_positions = model["comp_positions"]
    comp_of = model["comp_of"]
    over_at = model["over_at"]
    n = len(crossings)
    C = len(comp_positions)

    visit_lengths = [len(cp) for cp in comp_positions]
    strand_cv = _cv(visit_lengths)
    strand_entropy = _norm_entropy(visit_lengths)
    strand_ratio = (max(visit_lengths) / min(visit_lengths)) if min(visit_lengths) else float("inf")

    L = np.zeros((C, C), int)
    n_self = 0
    for cr in crossings:
        ci = comp_of[cr["odd"]]
        cj = comp_of[cr["even"]]
        if ci == cj:
            n_self += 1
            L[ci, ci] += 1
        else:
            L[ci, cj] += 1
            L[cj, ci] += 1
    n_inter = n - n_self
    link_deg = [int(np.sum(L[i]) - L[i, i]) for i in range(C)]
    link_deg_cv = _cv(link_deg) if C > 1 else 0.0

    signs = [1 if cr["even_signed"] > 0 else -1 for cr in crossings]
    n_pos = sum(1 for s in signs if s > 0)
    n_neg = n - n_pos
    sign_imbalance = abs(n_pos - n_neg) / n

    flips = total = 0
    for cp in comp_positions:
        Lc = len(cp)
        if Lc < 2:
            continue
        for i in range(Lc):
            total += 1
            if over_at[cp[i]] != over_at[cp[(i + 1) % Lc]]:
                flips += 1
    alternating_frac = (flips / total) if total else 1.0

    return {
        "n_crossings": n, "n_components": C,
        "strand_visit_lengths": visit_lengths, "strand_length_cv": strand_cv,
        "strand_balance_entropy": strand_entropy, "strand_length_ratio": strand_ratio,
        "n_self_crossings": n_self, "n_inter_crossings": n_inter,
        "linking_matrix": L.tolist(), "linking_degrees": link_deg,
        "linking_degree_cv": link_deg_cv,
        "n_over_neg_convention_pos": n_pos, "n_neg": n_neg,
        "sign_imbalance": sign_imbalance, "alternating_fraction": alternating_frac,
    }


def _is_seg(node):
    return isinstance(node, tuple) and len(node) == 2 and node[0] == "seg"


def _crossing_of_corner(node):
    if isinstance(node, tuple) and not _is_seg(node) and isinstance(node[0], int):
        return node[0]
    return None


def planar_graph_metrics(model, G):
    ok, emb = nx.check_planarity(G)
    if not ok:
        raise RuntimeError("Diagram graph is not planar (DT may be non-realizable).")
    n = len(model["crossings"])
    faces = DDOL.planar_faces(emb)
    face_degrees = []
    for f in faces:
        segs = [x for x in f if _is_seg(x)]
        if segs:
            face_degrees.append(len(segs))
    face_degrees.sort()
    n_bigons = sum(1 for d in face_degrees if d == 2)
    return {
        "euler_faces": len(face_degrees), "euler_faces_expected": n + 2,
        "face_degrees": face_degrees,
        "face_degree_cv": _cv(face_degrees) if face_degrees else 0.0,
        "n_bigons": n_bigons,
        "automorphism_order": None,   # injected by score_diagram (robust canonical count)
        "_emb": emb,
    }


def _signed_automorphism_order(model, G, cap=20000, time_limit=5.0):
    """Order of the diagram's symmetry group that PRESERVES over/under (the meaningful
    'same sequence' symmetry).  Nodes are labelled O/U/S and further coloured by
    Weisfeiler-Lehman refinement so VF2 only tries colour-preserving maps (automorphisms
    always preserve those colours, so the count is exact); a wall-clock guard in the main
    thread prevents pathological VF2 backtracking from hanging."""
    import threading
    H = nx.Graph(G)
    over_at = model["over_at"]
    labels = {}
    for k, cr in enumerate(model["crossings"]):
        oo, oe = bool(over_at[cr["odd"]]), bool(over_at[cr["even"]])
        for role in ("in_o", "out_o"):
            labels[(k, role)] = "O" if oo else "U"
        for role in ("in_e", "out_e"):
            labels[(k, role)] = "O" if oe else "U"
    for node in H.nodes():
        labels.setdefault(node, "S")
    try:
        wl = nx.weisfeiler_lehman_subgraph_hashes(H, iterations=4)
        colors = {node: "%s|%s" % (labels[node], (h[-1] if h else "")) for node, h in wl.items()}
    except Exception:
        colors = labels
    nx.set_node_attributes(H, colors, "_c")
    nm = nx.algorithms.isomorphism.categorical_node_match("_c", None)
    GM = nx.algorithms.isomorphism.GraphMatcher(H, H, node_match=nm)
    count = {"n": 0}

    def _run():
        for _ in GM.isomorphisms_iter():
            count["n"] += 1
            if count["n"] >= cap:
                break

    timed_out = False
    if threading.current_thread() is threading.main_thread():
        import signal

        def _handler(signum, frame):
            raise TimeoutError()

        old = signal.signal(signal.SIGALRM, _handler)
        signal.setitimer(signal.ITIMER_REAL, time_limit)
        try:
            _run()
        except TimeoutError:
            timed_out = True
        finally:
            signal.setitimer(signal.ITIMER_REAL, 0)
            signal.signal(signal.SIGALRM, old)
    else:
        _run()
    return count["n"], count["n"] >= cap, timed_out


def _crossing_centers_2d(model, P):
    n = len(model["crossings"])
    centers = np.zeros((n, 2), float)
    counts = np.zeros(n, float)
    for node, xy in P.items():
        k = _crossing_of_corner(node)
        if k is not None:
            centers[k] += np.asarray(xy, float)
            counts[k] += 1
    counts[counts == 0] = 1
    return centers / counts[:, None]


def _diagram_arcs(model):
    pos_cross = model["pos_cross"]
    arcs = []
    for cp in model["comp_positions"]:
        Lc = len(cp)
        for i in range(Lc):
            arcs.append((pos_cross[cp[i]], pos_cross[cp[(i + 1) % Lc]]))
    return arcs


def _best_rotational_symmetry_2d(centers, kmax=8, thresh=0.12):
    X = np.asarray(centers, float)
    X = X - X.mean(axis=0)
    scale = float(np.sqrt(np.mean(np.sum(X ** 2, axis=1)))) + 1e-15
    best_k, best_score = 1, 0.0
    for k in range(2, kmax + 1):
        Y = X @ _rotation_2d(2 * math.pi / k).T
        resid = _assignment_residual(X, Y) / scale
        score = max(0.0, 1.0 - resid)
        if resid < thresh and k > best_k:
            best_k = k
        best_score = max(best_score, score) if k == 2 else best_score
    if best_k > 1:
        Y = X @ _rotation_2d(2 * math.pi / best_k).T
        best_score = max(0.0, 1.0 - _assignment_residual(X, Y) / scale)
    return best_k, best_score


def geometric_2d_metrics(model, G):
    P = DDOL.compute_positions(G, "tutte")
    centers = _crossing_centers_2d(model, P)
    span = float(np.sqrt(np.mean(np.sum((centers - centers.mean(0)) ** 2, axis=1)))) + 1e-15
    C = centers / span
    arcs = _diagram_arcs(model)
    lengths = np.array([np.linalg.norm(C[a] - C[b]) for a, b in arcs], float)
    edge_cv = _cv(lengths)
    dirichlet_norm = float(np.mean(lengths ** 2) / (np.mean(lengths) ** 2 + 1e-15))
    from collections import defaultdict
    nbrs = defaultdict(list)
    for a, b in arcs:
        nbrs[a].append(b)
        nbrs[b].append(a)
    devs = []
    for k in range(len(C)):
        others = nbrs.get(k, [])
        if len(others) < 2:
            continue
        angs = sorted(math.atan2(*(C[o] - C[k])[::-1]) for o in others)
        gaps = [(angs[(i + 1) % len(angs)] - angs[i]) % (2 * math.pi) for i in range(len(angs))]
        ideal = 2 * math.pi / len(gaps)
        devs.append(math.degrees(math.sqrt(np.mean([(g - ideal) ** 2 for g in gaps]))))
    angle_dev = float(np.mean(devs)) if devs else 0.0
    sym_k, sym_score = _best_rotational_symmetry_2d(centers)
    return {
        "edge_length_cv": edge_cv, "dirichlet_energy_norm": dirichlet_norm,
        "crossing_angle_rms_dev_deg": angle_dev,
        "sym2d_order": sym_k, "sym2d_score": sym_score,
        "_centers2d": centers, "_arcs": arcs,
    }


def _crossing_centers_3d(model, G):
    dirs = DDOL._kamada_3d_unit_directions(G)
    n = len(model["crossings"])
    centers = np.zeros((n, 3), float)
    counts = np.zeros(n, float)
    for node, d in dirs.items():
        k = _crossing_of_corner(node)
        if k is not None:
            centers[k] += np.asarray(d, float)
            counts[k] += 1
    counts[counts == 0] = 1
    centers = centers / counts[:, None]
    norms = np.linalg.norm(centers, axis=1, keepdims=True)
    norms[norms < 1e-12] = 1.0
    return centers / norms


def _riesz_energy(points, s=1.0):
    P = np.asarray(points, float)
    m = len(P)
    e = 0.0
    for i in range(m):
        for j in range(i + 1, m):
            d = np.linalg.norm(P[i] - P[j])
            if d > 1e-12:
                e += 1.0 / d ** s
    return float(e)


def _best_rotational_symmetry_3d(centers, kmax=6, thresh=0.12):
    X = np.asarray(centers, float)
    X = X - X.mean(axis=0)
    scale = float(np.sqrt(np.mean(np.sum(X ** 2, axis=1)))) + 1e-15
    _, _, Vt = np.linalg.svd(X - X.mean(0), full_matrices=False)
    axes = [Vt[i] for i in range(Vt.shape[0])]
    cen = X.mean(0)
    for i in range(len(X)):
        v = X[i] - cen
        if np.linalg.norm(v) > 1e-9:
            axes.append(v)
    best_k, best_score = 1, 0.0
    for axis in axes:
        for k in range(2, kmax + 1):
            Y = X @ _rotation_about_axis(axis, 2 * math.pi / k).T
            resid = _assignment_residual(X, Y) / scale
            score = max(0.0, 1.0 - resid)
            if resid < thresh and k > best_k:
                best_k, best_score = k, score
            elif k == best_k and score > best_score:
                best_score = score
    if best_k == 1:
        best_score = 0.0
    return best_k, best_score


def sphere_3d_metrics(model, G):
    centers = _crossing_centers_3d(model, G)
    n = len(centers)
    thomson = _riesz_energy(centers, s=1.0)
    ref = _riesz_energy(DDOL._fibonacci_sphere_directions(n), s=1.0)
    spread_quality = float(ref / thomson) if thomson > 0 else 1.0
    pos_cross = model["pos_cross"]
    comp_lengths, turning = [], []
    for cp in model["comp_positions"]:
        ks = [pos_cross[p] for p in cp]
        pts = centers[ks]
        m = len(pts)
        if m < 2:
            comp_lengths.append(0.0)
            continue
        clen = 0.0
        for i in range(m):
            clen += math.acos(float(np.clip(np.dot(pts[i], pts[(i + 1) % m]), -1, 1)))
        comp_lengths.append(clen)
        for i in range(m):
            t1 = pts[i] - pts[(i - 1) % m]
            t2 = pts[(i + 1) % m] - pts[i]
            n1, n2 = np.linalg.norm(t1), np.linalg.norm(t2)
            if n1 > 1e-9 and n2 > 1e-9:
                ang = math.acos(float(np.clip(np.dot(t1, t2) / (n1 * n2), -1, 1)))
                turning.append(ang ** 2)
    strand3d_cv = _cv(comp_lengths)
    bending = float(np.sum(turning))
    sym_k, sym_score = _best_rotational_symmetry_3d(centers)
    return {
        "thomson_energy": thomson, "thomson_reference": ref,
        "sphere_spread_quality": spread_quality, "strand3d_length_cv": strand3d_cv,
        "bending_energy": bending, "sym3d_order": sym_k, "sym3d_score": sym_score,
        "_centers3d": centers,
    }


# Edit WEIGHTS to explore which properties should dominate the composite score.
WEIGHTS = {
    "strand_balance": 1.0,     # equal strand lengths across components
    "diagram_symmetry": 1.0,   # combinatorial + geometric symmetry
    "geometric_strain": 1.0,   # uniform arcs, ~90 deg crossings, regular faces
    "sphere_energy": 1.0,      # evenly spread crossings + low bending in 3-D
}


def _quality_scores(m):
    c, g, q2, q3 = m["combinatorial"], m["graph"], m["geom2d"], m["sphere3d"]
    strand_balance = 0.5 * c["strand_balance_entropy"] + 0.5 * (1.0 / (1.0 + c["strand_length_cv"]))
    aut = g["automorphism_order"]
    sym_comb = 1.0 - 1.0 / max(1, aut)
    diagram_symmetry = np.mean([sym_comb, q2["sym2d_score"], q3["sym3d_score"]])
    geometric_strain = np.mean([
        1.0 / (1.0 + q2["edge_length_cv"]),
        1.0 - min(1.0, q2["crossing_angle_rms_dev_deg"] / 90.0),
        1.0 / (1.0 + g["face_degree_cv"]),
    ])
    sphere_energy = np.mean([q3["sphere_spread_quality"], 1.0 / (1.0 + q3["strand3d_length_cv"])])
    return {
        "strand_balance": float(strand_balance),
        "diagram_symmetry": float(diagram_symmetry),
        "geometric_strain": float(geometric_strain),
        "sphere_energy": float(sphere_energy),
    }


def score_diagram(dt_string, negative_even="over"):
    comps = DDOL.parse_dt(dt_string)
    model = DDOL.build_model(comps, negative_even=negative_even)
    G = DDOL.build_gadget_graph(model)
    m = {
        "dt": dt_string,
        "combinatorial": combinatorial_metrics(model),
        "graph": planar_graph_metrics(model, G),
        "geom2d": geometric_2d_metrics(model, G),
        "sphere3d": sphere_3d_metrics(model, G),
    }
    # robust, reproducible symmetry order (count of DT relabellings fixing the canonical
    # code); replaces the fragile VF2 automorphism enumeration.
    m["graph"]["automorphism_order"] = canonical_symmetry(dt_string)
    m["quality"] = _quality_scores(m)
    m["composite"] = float(sum(WEIGHTS[k] * m["quality"][k] for k in WEIGHTS) / sum(WEIGHTS.values()))
    return m


def _strip_private(m):
    out = {}
    for k, v in m.items():
        if isinstance(v, dict):
            out[k] = {kk: vv for kk, vv in v.items() if not kk.startswith("_")}
        else:
            out[k] = v
    return out


# --------------------------------------------------------------------------- #
#  1. Generation  (SnapPy global + backtrack, re-rooted each round)
# --------------------------------------------------------------------------- #
def simplify_once(dt, backtrack_rounds, backtrack_steps, seed):
    """One round: SnapPy global simplify with backtrack, return the new DT string."""
    import snappy
    random.seed(seed)
    try:
        np.random.seed(seed & 0x7FFFFFFF)
    except Exception:
        pass
    L = snappy.Link(dt)
    L = LE.backtrack_simplify(snappy, L, mode="global",
                              rounds=backtrack_rounds, steps=backtrack_steps)
    return LE.dt_to_string(LE.parse_dt_any(L.DT_code()))


def _read_checkpoint(path):
    chain = {}
    if path and os.path.exists(path):
        with open(path) as fh:
            for line in fh:
                line = line.strip()
                if not line:
                    continue
                rec = json.loads(line)
                chain[int(rec["round"])] = rec["dt"]
    return [chain[i] for i in sorted(chain)] if chain else []


def _append_checkpoint(path, rnd, dt):
    if not path:
        return
    with open(path, "a") as fh:
        fh.write(json.dumps({"round": rnd, "dt": dt}) + "\n")


def generate_chain(dt0, rounds, backtrack_rounds, backtrack_steps, base_seed,
                   checkpoint=None, max_seconds=None, reset_every=0, verbose=True):
    """Return the list of DT strings [dt0, dt1, ..., dt_rounds]; resumable via checkpoint.

    If ``reset_every > 0`` the chain is re-rooted at ``dt0`` after every
    ``reset_every`` rounds (i.e. rounds reset_every+1, 2*reset_every+1, ...
    start again from the original diagram).  This prevents the walk from getting
    trapped cycling among a few common minimal diagrams and re-seeds exploration
    from the canonical starting point."""
    chain = _read_checkpoint(checkpoint)
    if not chain:
        chain = [dt0]
        _append_checkpoint(checkpoint, 0, dt0)
    elif chain[0] != dt0:
        raise ValueError("Checkpoint root DT does not match --dt; use a fresh --checkpoint.")

    t0 = time.time()
    while (len(chain) - 1) < rounds:
        i = len(chain)                       # next round index (1..rounds)
        seed = (int(base_seed) * 1000003 + i) & 0x7FFFFFFF
        root = chain[-1]
        if reset_every and i > 1 and ((i - 1) % int(reset_every) == 0):
            root = dt0                       # periodic re-root at the original diagram
        dt_new = simplify_once(root, backtrack_rounds, backtrack_steps, seed)
        chain.append(dt_new)
        _append_checkpoint(checkpoint, i, dt_new)
        if verbose and (i % 20 == 0 or i == rounds):
            print("  round %4d/%d  %.1fs" % (i, rounds, time.time() - t0), flush=True)
        if max_seconds and (time.time() - t0) >= max_seconds:
            break
    return chain


# --------------------------------------------------------------------------- #
#  2. Dedup  (exact signed-diagram isomorphism)
# --------------------------------------------------------------------------- #
def _iso_graph(dt):
    """Gadget graph with node labels encoding over/under; abstract-iso of this
    graph == same diagram up to rotation / reflection / relabelling / cyclic
    permutation / component reorder, while preserving the over/under pattern."""
    comps = DDOL.parse_dt(dt)
    model = DDOL.build_model(comps)
    G = DDOL.build_gadget_graph(model)
    over_at = model["over_at"]
    labels = {}
    for k, cr in enumerate(model["crossings"]):
        over_o = bool(over_at[cr["odd"]])     # is the odd strand over here?
        over_e = bool(over_at[cr["even"]])
        for role in ("in_o", "out_o"):
            labels[(k, role)] = "O" if over_o else "U"
        for role in ("in_e", "out_e"):
            labels[(k, role)] = "O" if over_e else "U"
    for node in G.nodes():
        if node not in labels:
            labels[node] = "S"                # traversal-arc (segment) node
    nx.set_node_attributes(G, labels, "lab")
    return G, model


def canonical_key(dt):
    """Strong, cheap composite signature of a signed diagram:
    (Weisfeiler-Lehman hash of the over/under-labelled diagram graph,
     strand-length spectrum, planar face-degree spectrum).
    Isomorphic diagrams (rotation/reflection/relabel/cyclic-perm/component-reorder,
    over-under preserved) share this signature; collisions between genuinely
    different diagrams are astronomically unlikely, so no per-member VF2 is needed."""
    G, model = _iso_graph(dt)
    wl = nx.weisfeiler_lehman_graph_hash(G, node_attr="lab", iterations=5)
    strand = tuple(sorted(len(cp) for cp in model["comp_positions"]))
    ok, emb = nx.check_planarity(G)
    fdeg = []
    if ok:
        for f in DDOL.planar_faces(emb):
            segs = [x for x in f if _is_seg(x)]
            if segs:
                fdeg.append(len(segs))
    return (wl, strand, tuple(sorted(fdeg)))


def _exact_iso(dtA, dtB):
    GA, _ = _iso_graph(dtA)
    GB, _ = _iso_graph(dtB)
    if GA.number_of_nodes() != GB.number_of_nodes():
        return False
    nm = nx.algorithms.isomorphism.categorical_node_match("lab", "")
    return nx.is_isomorphic(GA, GB, node_match=nm)


# --------------------------------------------------------------------------- #
#  Rigorous DT-native canonical form (authoritative "same diagram" test)
# --------------------------------------------------------------------------- #
# Two DT codes describe the SAME diagram iff, after re-deriving the DT under every
# choice of (component order, per-component base point, per-component traversal
# direction) [and optionally mirror = swapping over/under everywhere], their
# lexicographically-smallest valid DT codes are equal.  This is the exact meaning
# of "redundant due to symmetry / rotation / flipping / cyclic permutation", stated
# natively in DT terms and verifiable by hand.  It is O(product of component lengths)
# per diagram, so it is used for verification, not for bulk dedup.
def _diagram_tours(dt):
    m = DDOL.build_model(DDOL.parse_dt(dt))
    tours = [[(m["pos_cross"][p], bool(m["over_at"][p])) for p in cp]
             for cp in m["comp_positions"]]
    return tours, len(m["crossings"])


def _component_variants(tour, flip):
    base = [(c, (not o) if flip else o) for c, o in tour]
    L = len(base)
    out = []
    for seq in (base, base[::-1]):
        for s in range(L):
            out.append(seq[s:] + seq[:s])
    return out


def _walk_to_dt(walk, n, comp_bounds):
    slots = [[] for _ in range(n)]
    for i, (cr, ov) in enumerate(walk):
        slots[cr].append((i + 1, ov))
    signed = {}
    for lst in slots:
        if len(lst) != 2:
            return None
        (p1, o1), (p2, o2) = lst
        if (p1 & 1) == (p2 & 1):
            return None                        # not a valid DT for this base point
        if p1 & 1:
            oddp, evenp, ev = p1, p2, o2
        else:
            oddp, evenp, ev = p2, p1, o1
        signed[oddp] = (-evenp if ev else evenp)   # convention: negative even = even over
    tup = []
    for lo, hi in comp_bounds:
        odds = sorted(p for p in signed if (p & 1) and lo < p <= hi)
        tup.append(tuple(signed[p] for p in odds))
    return tuple(tup)


def canonical_dt(dt, allow_flip=True, return_symmetry=False):
    """Lexicographically minimal valid signed DT over all relabellings (and, if
    allow_flip, mirror).  Equal canonical_dt  <=>  same diagram.  When return_symmetry
    is True, also return how many valid relabellings reproduce that minimal code, which
    is the diagram's symmetry order (the count of DT re-encodings that coincide with the
    canonical form) - a robust, VF2-free symmetry measure."""
    from itertools import permutations, product
    tours, n = _diagram_tours(dt)
    C = len(tours)
    best = None
    n_min = 0
    for flip in ((False, True) if allow_flip else (False,)):
        var = [_component_variants(tours[ci], flip) for ci in range(C)]
        for perm in permutations(range(C)):
            bounds, off = [], 0
            for ci in perm:
                bounds.append((off, off + len(tours[ci])))
                off += len(tours[ci])
            for choice in product(*[var[ci] for ci in perm]):
                walk = []
                for seq in choice:
                    walk.extend(seq)
                tup = _walk_to_dt(walk, n, bounds)
                if tup is None:
                    continue
                if best is None or tup < best:
                    best, n_min = tup, 1
                elif tup == best:
                    n_min += 1
    return (best, n_min) if return_symmetry else best


def dedup(chain):
    """Collapse identical diagrams.  Returns a list of class dicts sorted by first
    appearance.  Groups by the composite signature from canonical_key(), computed once
    per *distinct* DT string (fast: no per-member graph isomorphism search)."""
    # 1. group member round-indices by exact DT string, preserving first-seen order
    str_members, str_first = {}, {}
    for idx, dt in enumerate(chain):
        if dt not in str_members:
            str_members[dt] = []
            str_first[dt] = idx
        str_members[dt].append(idx)
    unique = sorted(str_members, key=lambda s: str_first[s])

    # 2. signature once per distinct string; group by signature.
    #    The signature (WL hash of the over/under-labelled diagram graph + strand-length
    #    spectrum + face-degree spectrum) is an isomorphism invariant, so identical
    #    diagrams always land together; a collision between genuinely different diagrams
    #    is astronomically unlikely and can be ruled out with --verify (exact VF2) or
    #    canonical_dt().
    groups = {}
    str_strings = {}
    for dt in unique:
        sig = canonical_key(dt)
        groups.setdefault(sig, []).append(dt)

    classes = []
    for sig, strings in groups.items():
        members = sorted(m for s in strings for m in str_members[s])
        classes.append({
            "rep_dt": strings[0],              # earliest first-seen string in the class
            "strings": strings,
            "members": members,
            "multiplicity": len(members),
            "rep_round": min(members),
            "n_distinct_strings": len(strings),
            "sig": sig,
        })
    classes.sort(key=lambda c: c["rep_round"])
    for j, c in enumerate(classes, start=1):
        c["rep_id"] = j
    return classes


def verify_classes(classes, sample=25):
    """Confidence check: within each class, run exact labelled-graph VF2 between the
    representative and up to `sample` other distinct DT strings (0 = all).  Returns a
    per-class report; a False means the fast signature merged non-isomorphic diagrams."""
    import random as _r
    nm = nx.algorithms.isomorphism.categorical_node_match("lab", "")
    report = []
    for c in classes:
        others = [s for s in c["strings"] if s != c["rep_dt"]]
        if sample and len(others) > sample:
            others = _r.Random(0).sample(others, sample)
        Grep, _ = _iso_graph(c["rep_dt"])
        ok = all(nx.is_isomorphic(Grep, _iso_graph(o)[0], node_match=nm) for o in others)
        report.append({"rep_id": c["rep_id"], "checked": len(others), "all_isomorphic": ok})
    return report


def check_sampled(classes, queries):
    """For each query DT code, report whether an equivalent diagram was sampled
    (i.e. its signature matches one of the representative classes)."""
    sigmap = {c["sig"]: c for c in classes}
    out = []
    for q in queries:
        try:
            k = canonical_key(q)
        except Exception as exc:  # noqa: BLE001
            out.append({"dt": q, "sampled": False, "error": str(exc)})
            continue
        c = sigmap.get(k)
        out.append({
            "dt": q,
            "sampled": c is not None,
            "matches_rep_id": (c["rep_id"] if c else None),
            "multiplicity": (c["multiplicity"] if c else 0),
            "first_round": (c["rep_round"] if c else None),
        })
    return out


# --------------------------------------------------------------------------- #
#  3. Score + rank
# --------------------------------------------------------------------------- #
def _jones(dt):
    try:
        import snappy
        return str(snappy.Link(dt).jones_polynomial())
    except Exception:
        return "n/a (needs Sage)"


def _linking_fp(dt):
    """Headless same-link fingerprint: sorted off-diagonal linking numbers."""
    try:
        import snappy
        M = snappy.Link(dt).linking_matrix()
        vals = sorted(int(M[i][j]) for i in range(len(M)) for j in range(len(M)) if i < j)
        return tuple(vals)
    except Exception:
        return None


_CANON_CACHE = None
_CANON_CACHE_PATH = os.environ.get("CANON_CACHE", "canonical_cache.json")


def _canonical_entry(dt, allow_flip=False):
    """Return {'dt': canonical string, 'sym': symmetry order}, memoized on disk so the
    (somewhat expensive) canonicalization is computed once per diagram per run."""
    global _CANON_CACHE
    if _CANON_CACHE is None:
        _CANON_CACHE = {}
        if os.path.exists(_CANON_CACHE_PATH):
            try:
                with open(_CANON_CACHE_PATH) as fh:
                    _CANON_CACHE = json.load(fh)
            except Exception:
                _CANON_CACHE = {}
    key = ("F:" if allow_flip else "N:") + dt
    entry = _CANON_CACHE.get(key)
    if isinstance(entry, dict) and "dt" in entry and "sym" in entry:
        return entry
    t, n_min = canonical_dt(dt, allow_flip=allow_flip, return_symmetry=True)
    s = "DT: [" + ", ".join("(" + ", ".join(str(x) for x in comp) + ")" for comp in t) + "]"
    entry = {"dt": s, "sym": int(n_min)}
    _CANON_CACHE[key] = entry
    _CANON_CACHE[("F:" if allow_flip else "N:") + s] = entry   # canonical maps to itself
    try:
        with open(_CANON_CACHE_PATH, "w") as fh:
            json.dump(_CANON_CACHE, fh)
    except Exception:
        pass
    return entry


def canonical_dt_string(dt, allow_flip=False):
    return _canonical_entry(dt, allow_flip)["dt"]


def canonical_symmetry(dt, allow_flip=False):
    return _canonical_entry(dt, allow_flip)["sym"]


def score_representatives(classes):
    """Score each class on its CANONICAL DT code so the geometric/energy metrics are
    reproducible and independent of which sampled DT string first represented the class
    (the layout of two isomorphic-but-differently-labelled codes can differ, which would
    otherwise make the composite score wobble between runs)."""
    scored = []
    for c in classes:
        cdt = canonical_dt_string(c["rep_dt"])
        c["canonical_dt"] = cdt
        m = score_diagram(cdt)            # m["dt"] == canonical form
        m["_class"] = c
        m["jones"] = _jones(cdt)
        m["linking_fp"] = _linking_fp(cdt)
        scored.append(m)
    scored.sort(key=lambda m: m["composite"], reverse=True)
    for rank, m in enumerate(scored, start=1):
        m["rank"] = rank
    return scored


# --------------------------------------------------------------------------- #
#  4a. Excel report
# --------------------------------------------------------------------------- #
def _cols():
    # (header, path-getter) ; getter takes the scored metrics dict m
    def g(*keys):
        def _get(m):
            d = m
            for k in keys:
                d = d[k]
            return d
        return _get
    return [
        ("Rank", lambda m: m["rank"]),
        ("Rep ID", lambda m: m["_class"]["rep_id"]),
        ("DT code", lambda m: m["dt"]),
        ("Multiplicity", lambda m: m["_class"]["multiplicity"]),
        ("First round", lambda m: m["_class"]["rep_round"]),
        ("Crossings", g("combinatorial", "n_crossings")),
        ("Components", g("combinatorial", "n_components")),
        ("Strand lengths", lambda m: str(m["combinatorial"]["strand_visit_lengths"])),
        ("Strand length CV", g("combinatorial", "strand_length_cv")),
        ("Strand balance entropy", g("combinatorial", "strand_balance_entropy")),
        ("Strand length max/min", g("combinatorial", "strand_length_ratio")),
        ("Self crossings", g("combinatorial", "n_self_crossings")),
        ("Inter crossings", g("combinatorial", "n_inter_crossings")),
        ("Linking degree CV", g("combinatorial", "linking_degree_cv")),
        ("Alternating fraction", g("combinatorial", "alternating_fraction")),
        ("Over/under imbalance", g("combinatorial", "sign_imbalance")),
        ("Faces (n+2)", g("graph", "euler_faces")),
        ("Face degree CV", g("graph", "face_degree_cv")),
        ("Bigons", g("graph", "n_bigons")),
        ("Symmetry order", g("graph", "automorphism_order")),
        ("Edge length CV", g("geom2d", "edge_length_cv")),
        ("Dirichlet energy", g("geom2d", "dirichlet_energy_norm")),
        ("Crossing angle dev (deg)", g("geom2d", "crossing_angle_rms_dev_deg")),
        ("2D symmetry score", g("geom2d", "sym2d_score")),
        ("Thomson energy", g("sphere3d", "thomson_energy")),
        ("Sphere spread quality", g("sphere3d", "sphere_spread_quality")),
        ("3D strand length CV", g("sphere3d", "strand3d_length_cv")),
        ("Bending energy", g("sphere3d", "bending_energy")),
        ("3D symmetry order", g("sphere3d", "sym3d_order")),
        ("q: strand balance", g("quality", "strand_balance")),
        ("q: diagram symmetry", g("quality", "diagram_symmetry")),
        ("q: geometric strain", g("quality", "geometric_strain")),
        ("q: sphere energy", g("quality", "sphere_energy")),
        ("COMPOSITE", lambda m: m["composite"]),
        ("Linking numbers (sorted)", lambda m: str(m.get("linking_fp"))),
        ("Jones (Sage only)", lambda m: m["jones"]),
    ]


def write_excel(scored, path, run_info):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    cols = _cols()
    wb = Workbook()
    ws = wb.active
    ws.title = "representatives"

    head_fill = PatternFill("solid", fgColor="1F3864")
    head_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    best_fill = PatternFill("solid", fgColor="C6EFCE")
    cell_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cidx, (header, _) in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=cidx, value=header)
        cell.fill = head_fill
        cell.font = head_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for ridx, m in enumerate(scored, start=2):
        best = (m["rank"] == 1)
        for cidx, (_, getter) in enumerate(cols, start=1):
            val = getter(m)
            if isinstance(val, float):
                val = round(val, 4)
            cell = ws.cell(row=ridx, column=cidx, value=val)
            cell.font = cell_font
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if best:
                cell.fill = best_fill

    # widths
    for cidx, (header, _) in enumerate(cols, start=1):
        letter = get_column_letter(cidx)
        if header == "DT code":
            ws.column_dimensions[letter].width = 42
        elif header in ("Jones (same link check)", "Strand lengths"):
            ws.column_dimensions[letter].width = 22
        else:
            ws.column_dimensions[letter].width = max(11, min(20, len(header) + 2))
    ws.freeze_panes = "D2"
    ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(cols)), len(scored) + 1)

    # run info sheet
    ws2 = wb.create_sheet("run_info")
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 80
    for r, (k, v) in enumerate(run_info.items(), start=1):
        a = ws2.cell(row=r, column=1, value=k)
        a.font = Font(name="Arial", bold=True, size=10)
        b = ws2.cell(row=r, column=2, value=str(v))
        b.font = Font(name="Arial", size=10)
        b.alignment = Alignment(wrap_text=True, vertical="top")

    # legend sheet: for every metric column, whether higher or lower is better
    ws3 = wb.create_sheet("metric_legend")
    for col, w in (("A", 26), ("B", 20), ("C", 74)):
        ws3.column_dimensions[col].width = w
    green = PatternFill("solid", fgColor="C6EFCE")   # higher = better
    blue = PatternFill("solid", fgColor="DDEBF7")    # lower = better
    grey = PatternFill("solid", fgColor="EDEDED")    # descriptive / fixed
    for c, txt in enumerate(("Metric", "Direction", "Meaning"), start=1):
        cell = ws3.cell(row=1, column=c, value=txt)
        cell.fill = head_fill
        cell.font = head_font
        cell.border = border
    fillmap = {"higher = better": green, "lower = better": blue,
               "descriptive": grey, "fixed": grey}
    for r, (metric, direction, meaning) in enumerate(_METRIC_LEGEND, start=2):
        for c, val in enumerate((metric, direction, meaning), start=1):
            cell = ws3.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=(c == 3))
            if c == 2:
                cell.fill = fillmap.get(direction, grey)

    wb.save(path)


# For each Excel metric column: is larger better, smaller better, or just descriptive?
_METRIC_LEGEND = [
    ("Rank", "lower = better", "1 = best blueprint overall."),
    ("Multiplicity", "descriptive", "How often the simplifier produced this diagram; NOT a quality (the best diagram is actually rare)."),
    ("First round", "descriptive", "Round at which the diagram was first seen."),
    ("Crossings", "lower = better", "Diagram complexity; fixed at 14 here (all are minimal)."),
    ("Components", "fixed", "Number of strands; fixed at 4."),
    ("Strand length CV", "lower = better", "Spread of strand lengths; 0 = all strands equal length."),
    ("Strand balance entropy", "higher = better", "Evenness of strand lengths; 1 = perfectly even."),
    ("Strand length max/min", "lower = better", "Longest / shortest strand; 1 = equal."),
    ("Self crossings", "descriptive", "Within-strand crossings; 0 for this link."),
    ("Inter crossings", "descriptive", "Between-strand crossings."),
    ("Linking degree CV", "lower = better", "Evenness of how the linking load is shared; lower = more even."),
    ("Alternating fraction", "descriptive", "How close to an alternating diagram; not scored."),
    ("Over/under imbalance", "descriptive", "Convention-dependent over/under balance; not scored."),
    ("Faces (n+2)", "fixed", "Number of regions; equals crossings + 2 (Euler check) = 16."),
    ("Face degree CV", "lower = better", "Regularity of the regions; lower = more uniform tiling."),
    ("Bigons", "lower = better", "Two-sided regions (clasps); 0 = no local crowding."),
    ("Symmetry order", "higher = better", "How many ways the diagram maps onto itself; higher = fewer unique sequences."),
    ("Edge length CV", "lower = better", "Uniformity of segment lengths in the relaxed 2-D layout."),
    ("Dirichlet energy", "lower = better", "2-D spring energy; 1 = perfectly uniform springs."),
    ("Crossing angle dev (deg)", "lower = better", "Deviation of crossings from ideal 90-degree X shapes."),
    ("2D symmetry score", "higher = better", "Rotational symmetry of the laid-out crossings (0-1)."),
    ("Thomson energy", "lower = better", "Crowding of crossings on the sphere; lower = more evenly spread."),
    ("Sphere spread quality", "higher = better", "Evenness of the spread on the sphere vs ideal; 1 = ideal."),
    ("3D strand length CV", "lower = better", "Evenness of strand lengths measured on the 3-D sphere."),
    ("Bending energy", "lower = better", "How sharply strands turn in 3-D; lower = gentler, relaxed curves."),
    ("3D symmetry order", "higher = better", "Rotational symmetry axis of the 3-D layout (C-k)."),
    ("q: strand balance", "higher = better", "0-1 quality for even strands."),
    ("q: diagram symmetry", "higher = better", "0-1 quality for symmetry."),
    ("q: geometric strain", "higher = better", "0-1 quality for low 2-D strain."),
    ("q: sphere energy", "higher = better", "0-1 quality for even, relaxed 3-D layout."),
    ("COMPOSITE", "higher = better", "Overall 0-1 score; higher = better synthesis blueprint."),
    ("Linking numbers (sorted)", "descriptive", "Pairwise linking fingerprint (same for all: the link is fixed)."),
    ("Jones (Sage only)", "descriptive", "Link invariant; identical for all (same link). Populates under Sage."),
]


# --------------------------------------------------------------------------- #
#  4b. SVG figure  (2-D Tutte layout + 3-D sphere layout per representative)
# --------------------------------------------------------------------------- #
_PALETTE = ["#4c72b0", "#dd8452", "#55a868", "#c44e52", "#8172b3",
            "#937860", "#da8bc3", "#8c8c8c", "#ccb974", "#64b5cd"]


def _arc_list(model):
    """Directed traversal arcs (ci, a, b) between crossing indices, one per visit-step."""
    pos_cross = model["pos_cross"]
    arcs = []
    for ci, cp in enumerate(model["comp_positions"]):
        ks = [pos_cross[p] for p in cp]
        L = len(ks)
        for i in range(L):
            arcs.append((ci, ks[i], ks[(i + 1) % L]))
    return arcs


def _bow_offsets(arcs):
    """Offset multiplier for each arc so that arcs sharing the same endpoint pair
    (e.g. the two arcs of a bigon, or arcs of different components between the same
    crossings) fan apart instead of drawing on top of each other.  Single edges get 0."""
    from collections import defaultdict
    groups = defaultdict(list)
    for idx, (ci, a, b) in enumerate(arcs):
        groups[tuple(sorted((a, b)))].append(idx)
    off = {}
    for _, idxs in groups.items():
        k = len(idxs)
        for j, idx in enumerate(idxs):
            off[idx] = (j - (k - 1) / 2.0)     # symmetric around 0
    return off


def _draw_skeleton_2d(ax, C2, model, gap_frac=0.05):
    """2-D component-coloured skeleton; parallel arcs are bowed apart (quadratic Bezier)."""
    arcs = _arc_list(model)
    off = _bow_offsets(arcs)
    span = float(np.max(C2.max(0) - C2.min(0))) or 1.0
    gap = gap_frac * span
    ts = np.linspace(0, 1, 26)[:, None]
    for idx, (ci, a, b) in enumerate(arcs):
        pa, pb = C2[a], C2[b]
        col = _PALETTE[ci % len(_PALETTE)]
        # Use a CANONICAL direction (sorted endpoints) for the perpendicular so that two
        # parallel arcs between the same crossings bow to OPPOSITE sides regardless of the
        # direction each is traversed.  (Using each arc's own direction flips the perp and
        # made oppositely-traversed bigon arcs overlap.)
        lo, hi = (a, b) if a <= b else (b, a)
        dc = C2[hi] - C2[lo]
        L = float(np.hypot(dc[0], dc[1]))
        if L < 1e-9:
            continue
        perp = np.array([-dc[1], dc[0]]) / L
        ctrl = (pa + pb) / 2.0 + perp * (off[idx] * gap * 2.0)   # midpoint bow = off*gap
        pts = (1 - ts) ** 2 * pa + 2 * (1 - ts) * ts * ctrl + ts ** 2 * pb
        ax.plot(pts[:, 0], pts[:, 1], "-", lw=1.5, color=col, alpha=0.9)
    ax.plot(C2[:, 0], C2[:, 1], "o", ms=4.0, color="#222222", zorder=3)
    ax.set_aspect("equal")
    ax.axis("off")


def _draw_sphere_depth(ax3, C3, model, elev=22.0, azim=-58.0, gap=0.13):
    """3-D sphere layout: arcs ride on the sphere (renormalized), parallel arcs are
    bowed apart, and transparency is depth-cued (nearer = opaque, farther = faint)."""
    ax3.view_init(elev=elev, azim=azim)
    er, ar = math.radians(elev), math.radians(azim)
    eye = np.array([math.cos(er) * math.cos(ar),
                    math.cos(er) * math.sin(ar), math.sin(er)], float)
    u = np.linspace(0, 2 * np.pi, 26)
    v = np.linspace(0, np.pi, 13)
    ax3.plot_wireframe(np.outer(np.cos(u), np.sin(v)), np.outer(np.sin(u), np.sin(v)),
                       np.outer(np.ones_like(u), np.cos(v)), color="0.9", linewidth=0.2)
    arcs = _arc_list(model)
    off = _bow_offsets(arcs)
    ss = np.linspace(0, 1, 16)
    for idx, (ci, a, b) in enumerate(arcs):
        pa, pb = C3[a], C3[b]
        col = _PALETTE[ci % len(_PALETTE)]
        # canonical direction (sorted endpoints) so parallel arcs bow to opposite sides
        lo, hi = (a, b) if a <= b else (b, a)
        chord = C3[hi] - C3[lo]
        mid = (pa + pb) / 2.0
        t = np.cross(mid, chord)
        nt = np.linalg.norm(t)
        if nt < 1e-9:
            continue
        t = t / nt
        ctrl = mid + t * (off[idx] * gap * 2.0)
        # bowed arc, each sample renormalized to the unit sphere
        pts = ((1 - ss) ** 2)[:, None] * pa + (2 * (1 - ss) * ss)[:, None] * ctrl \
            + (ss ** 2)[:, None] * pb
        pts = pts / np.linalg.norm(pts, axis=1, keepdims=True)
        for j in range(len(ss) - 1):
            s0, s1 = pts[j], pts[j + 1]
            depth = float(np.dot((s0 + s1) / 2.0, eye))
            alpha = 0.12 + 0.83 * (depth + 1.0) / 2.0
            ax3.plot([s0[0], s1[0]], [s0[1], s1[1]], [s0[2], s1[2]],
                     "-", lw=1.7, color=col, alpha=alpha, solid_capstyle="round")
    dvals = C3 @ eye
    alphas = 0.2 + 0.8 * (dvals - dvals.min()) / (np.ptp(dvals) + 1e-9)
    for k in range(len(C3)):
        ax3.scatter(C3[k, 0], C3[k, 1], C3[k, 2], c="#222222",
                    s=9, alpha=float(alphas[k]), depthshade=False)
    ax3.set_box_aspect((1, 1, 1))
    ax3.axis("off")


# 2-D layout used for BOTH the draw_dt_original_labels panel and the Tutte skeleton,
# so their orientation/rotation is identical (per request: shaped-tutte, ellipse, aspect 1).
_DRAW_LAYOUT = "shaped-tutte"
_DRAW_TUTTE_OPTS = {"shape": "ellipse", "aspect": 1.0}
_DRAW_MIN_SEP = 0.02          # push apart non-incident strand pieces closer than this (fraction of span)
_SPHERE_VIEWS = [(22.0, -58.0), (22.0, 122.0)]   # two viewpoints (elev, azim), ~180 deg apart


def make_figure(scored, path, max_draw=60):
    import textwrap
    reps = scored[:max_draw]
    n = len(reps)
    total = sum(m["_class"]["multiplicity"] for m in scored)
    ncol = 3 + len(_SPHERE_VIEWS)   # text | draw | skeleton | sphere-view-1 | sphere-view-2
    # one representative per ROW
    fig = plt.figure(figsize=(3.0 * ncol + 5.0, 3.6 * n + 0.6))
    gs = fig.add_gridspec(n, ncol, width_ratios=[1.25, 1.3, 1.05] + [1.0] * len(_SPHERE_VIEWS),
                          hspace=0.28, wspace=0.05)

    for i, m in enumerate(reps):
        model = DDOL.build_model(DDOL.parse_dt(m["dt"]))
        G = DDOL.build_gadget_graph(model)
        # shared shaped-tutte layout -> identical rotation in the drawing and the skeleton
        P = DDOL.compute_positions(G, _DRAW_LAYOUT, tutte_opts=dict(_DRAW_TUTTE_OPTS))
        if _DRAW_MIN_SEP > 0:
            P = DDOL.nudge_min_separation(P, G, _DRAW_MIN_SEP)   # min-separation relaxation
        centers_d = DDOL.crossing_centers(model, P)
        C2 = np.array([centers_d[k] for k in range(len(model["crossings"]))], float)
        C3 = m["sphere3d"]["_centers3d"]
        rid = m["_class"]["rep_id"]
        col_of = lambda ci: _PALETTE[ci % len(_PALETTE)]

        # --- col 0: text (full DT code shown completely, wrapped) + metrics ---
        axt = fig.add_subplot(gs[i, 0])
        axt.axis("off")
        head = ("Rank %d   (rep #%d)\ncomposite = %.3f\nmultiplicity = %d / %d\n"
                "symmetry = %d   3D-sym = C%d\nstrands = %s\nbending = %.1f"
                % (m["rank"], rid, m["composite"], m["_class"]["multiplicity"], total,
                   m["graph"]["automorphism_order"], m["sphere3d"]["sym3d_order"],
                   m["combinatorial"]["strand_visit_lengths"],
                   m["sphere3d"]["bending_energy"]))
        dt_wrapped = "\n".join(textwrap.wrap(m["dt"], width=30))
        axt.text(0.0, 1.0, head, transform=axt.transAxes, ha="left", va="top",
                 fontsize=9.5, fontweight="bold" if m["rank"] == 1 else "normal")
        axt.text(0.0, 0.34, "DT code:", transform=axt.transAxes, ha="left", va="top",
                 fontsize=8.5, style="italic")
        axt.text(0.0, 0.27, dt_wrapped, transform=axt.transAxes, ha="left", va="top",
                 fontsize=8.0, family="monospace")

        # --- col 1: draw_dt_original_labels (shaped-tutte, ellipse, aspect 1, min-sep 0.02) ---
        axd = fig.add_subplot(gs[i, 1])
        try:
            DDOL.render_diagram(axd, model, P, centers_d, color_of=col_of,
                                show_labels=True, arrows=True, lw=1.6, label_fontsize=5.5)
        except Exception as exc:  # keep the grid robust
            axd.text(0.5, 0.5, "render error:\n%s" % exc, ha="center", va="center",
                     fontsize=6, transform=axd.transAxes)
        axd.set_aspect("equal")
        axd.axis("off")
        if i == 0:
            axd.set_title("draw_dt_original_labels\n(shaped-tutte ellipse, min-sep 0.02)", fontsize=9)

        # --- col 2: 2-D skeleton, SAME layout/rotation, parallel arcs bowed apart ---
        ax2 = fig.add_subplot(gs[i, 2])
        _draw_skeleton_2d(ax2, C2, model)
        if i == 0:
            ax2.set_title("2-D skeleton (same layout)", fontsize=9)

        # --- cols 3+: 3-D sphere layout from two viewpoints, depth-shaded ---
        for vi, (elev, azim) in enumerate(_SPHERE_VIEWS):
            ax3 = fig.add_subplot(gs[i, 3 + vi], projection="3d")
            _draw_sphere_depth(ax3, C3, model, elev=elev, azim=azim)
            if i == 0:
                ax3.set_title("3-D sphere (view %d)" % (vi + 1), fontsize=9)

    fig.suptitle("Representative diagrams of one link, ranked by composite score  "
                 "(full DT codes at left; 3-D sphere shown from two viewpoints)",
                 fontsize=13, y=0.999)
    fig.savefig(path, bbox_inches="tight")
    plt.close(fig)


# --------------------------------------------------------------------------- #
#  5. CLI
# --------------------------------------------------------------------------- #
DEFAULT_DT = "DT: [(-8,-12,16),(-24,-22,-28,-26),(-10,-14,-2),(-20,-6,-18,-4)]"


def launch_gui(defaults=None):
    """Tkinter front-end: fill in the parameters and press Run.  Launched when the
    script is started with no arguments or with --gui."""
    try:
        import tkinter as tk
        from tkinter import scrolledtext, filedialog
        root = tk.Tk()                       # fails here if there is no display
    except Exception as exc:  # no Tk / no display
        print("Tkinter GUI unavailable (%s)." % exc)
        print("Run with explicit CLI arguments such as --rounds, --xlsx, --svg, "
              "and --checkpoint to start a headless scoring job.")
        return

    import threading
    import queue as _queue

    def _apply_score_icon(window):
        if not os.path.exists(SCORE_ICON_PNG):
            return
        try:
            image = tk.PhotoImage(file=SCORE_ICON_PNG)
            window.iconphoto(True, image)
            window._score_diagram_icon = image
        except Exception:
            pass

    def dv(name, fallback):
        return str(getattr(defaults, name, fallback)) if defaults is not None else str(fallback)

    _apply_score_icon(root)
    root.title("DT Diagram Scorer  -  score_diagramV2_0")
    frm = tk.Frame(root, padx=10, pady=8)
    frm.pack(fill="x")
    frm.columnconfigure(1, weight=1)

    # GUI default for "Reset to root" is 20 (respects a non-zero value passed via --reset-every).
    reset_default = getattr(defaults, "reset_every", 0) or 20
    rows = [
        ("dt", "DT code", dv("dt", DEFAULT_DT)),
        ("rounds", "Rounds (simplifications)", dv("rounds", 99)),
        ("backtrack_rounds", "Backtrack rounds", dv("backtrack_rounds", 200)),
        ("backtrack_steps", "Backtrack steps", dv("backtrack_steps", 30)),
        ("reset_every", "Reset to root every N (0=off)", reset_default),
        ("seed", "Seed", dv("seed", 20260708)),
        ("checkpoint", "Checkpoint file", dv("checkpoint", "chainV2.jsonl")),
        ("max_draw", "Max diagrams drawn", dv("max_draw", 60)),
        ("verify", "Verify sample (0=off, -1=all)", dv("verify", 0)),
        ("xlsx", "Excel out (.xlsx, blank=skip)", getattr(defaults, "xlsx", "") or "diagram_scores.xlsx"),
        ("svg", "SVG figure out (blank=skip)", getattr(defaults, "svg", "") or "diagram_scores.svg"),
        ("json", "JSON out (blank=skip)", getattr(defaults, "json", "") or ""),
    ]
    # output-file rows get a "Browse..." button (choose folder + file name via a save dialog)
    _save_dialog = {
        "xlsx": [("Excel workbook", "*.xlsx")],
        "svg": [("SVG figure", "*.svg"), ("PNG image", "*.png")],
        "json": [("JSON", "*.json")],
    }

    def _make_browser(var, filetypes, defext):
        def _browse():
            path = filedialog.asksaveasfilename(
                title="Choose output location", defaultextension=defext,
                filetypes=filetypes + [("All files", "*.*")],
                initialfile=os.path.basename(var.get() or ""),
                initialdir=os.path.dirname(var.get() or "") or os.getcwd())
            if path:
                var.set(path)
        return _browse

    vars_ = {}
    for r, (key, label, val) in enumerate(rows):
        tk.Label(frm, text=label, anchor="w").grid(row=r, column=0, sticky="w", pady=2)
        v = tk.StringVar(value=str(val))
        tk.Entry(frm, textvariable=v, width=58).grid(row=r, column=1, sticky="we", padx=6, pady=2)
        if key in _save_dialog:
            ft = _save_dialog[key]
            tk.Button(frm, text="Browse…",
                      command=_make_browser(v, ft, ft[0][1].lstrip("*"))
                      ).grid(row=r, column=2, sticky="w", padx=(0, 2))
        vars_[key] = v

    # --- why the seed matters (explained inline) ---
    tk.Label(frm, anchor="w", justify="left", fg="#444444",
             text="Seed: the search is randomized — each round randomly tangles the diagram before\n"
                  "re-simplifying it, which is how it discovers different diagrams. The seed fixes that\n"
                  "randomness so the run is fully reproducible: the same seed always yields the same\n"
                  "sequence of diagrams and the same scores (and lets a run resume from its checkpoint\n"
                  "and be re-run identically for a paper). Change it to explore a different random path.")\
        .grid(row=len(rows), column=0, columnspan=3, sticky="w", pady=(8, 0))

    # --- Check DT codes: membership test (explained inline) ---
    tk.Label(frm, text="Check DT codes (optional, one DT per line):",
             anchor="w", font=("TkDefaultFont", 10, "bold"))\
        .grid(row=len(rows) + 1, column=0, columnspan=3, sticky="w", pady=(10, 0))
    tk.Label(frm, anchor="w", justify="left", fg="#444444",
             text="After the run, each DT code you paste here is tested against the diagrams that\n"
                  "were found: the log reports whether an equivalent diagram was sampled, and if so\n"
                  "which ranked representative it matches (and how often it occurred). Use it to ask\n"
                  "'is this particular diagram among the ones my search produced?'")\
        .grid(row=len(rows) + 2, column=0, columnspan=3, sticky="w")
    check_text = tk.Text(frm, width=64, height=4)
    check_text.grid(row=len(rows) + 3, column=0, columnspan=3, sticky="we", pady=2)

    btns = tk.Frame(root, padx=10, pady=4)
    btns.pack(fill="x")
    run_btn = tk.Button(btns, text="Run")
    run_btn.pack(side="left")
    tk.Button(btns, text="Quit", command=root.destroy).pack(side="right")

    log = scrolledtext.ScrolledText(root, width=104, height=20, font=("Menlo", 9))
    log.pack(fill="both", expand=True, padx=10, pady=(4, 10))

    q = _queue.Queue()

    class _QWriter:
        def write(self, s):
            q.put(s)

        def flush(self):
            pass

    def _poll():
        try:
            while True:
                log.insert("end", q.get_nowait())
                log.see("end")
        except _queue.Empty:
            pass
        root.after(120, _poll)

    def _run():
        try:
            a = argparse.Namespace(
                dt=vars_["dt"].get().strip() or DEFAULT_DT,
                rounds=int(vars_["rounds"].get()),
                backtrack_rounds=int(vars_["backtrack_rounds"].get()),
                backtrack_steps=int(vars_["backtrack_steps"].get()),
                reset_every=int(vars_["reset_every"].get()),
                seed=int(vars_["seed"].get()),
                checkpoint=vars_["checkpoint"].get().strip() or "chainV2.jsonl",
                max_seconds=0.0, generate_only=False,
                max_draw=int(vars_["max_draw"].get()),
                verify=int(vars_["verify"].get() or 0),
                xlsx=vars_["xlsx"].get().strip() or None,
                svg=vars_["svg"].get().strip() or None,
                json=vars_["json"].get().strip() or None,
                check=[ln.strip() for ln in check_text.get("1.0", "end").splitlines() if ln.strip()],
                check_file=None, gui=False,
            )
        except ValueError as exc:
            q.put("Invalid parameter: %s\n" % exc)
            return
        run_btn.config(state="disabled")
        q.put("\n===== run started =====\n")

        def _worker():
            old = sys.stdout
            sys.stdout = _QWriter()
            try:
                run_pipeline(a)
                print("\n===== done =====\n")
            except Exception:  # noqa: BLE001
                import traceback
                print("ERROR:\n" + traceback.format_exc())
            finally:
                sys.stdout = old
                root.after(0, lambda: run_btn.config(state="normal"))

        threading.Thread(target=_worker, daemon=True).start()

    run_btn.config(command=_run)
    q.put("Set parameters and press Run. Long runs stream progress here.\n"
          "Tip: a checkpoint file lets you stop and resume; large 'Rounds' can take minutes.\n")
    root.after(120, _poll)
    root.mainloop()


def main(argv=None):
    raw = list(sys.argv[1:]) if argv is None else list(argv)
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--dt", default=DEFAULT_DT)
    ap.add_argument("--rounds", type=int, default=99)
    ap.add_argument("--backtrack-rounds", type=int, default=200)
    ap.add_argument("--backtrack-steps", type=int, default=30)
    ap.add_argument("--seed", type=int, default=20260708)
    ap.add_argument("--checkpoint", default="chainV2.jsonl")
    ap.add_argument("--reset-every", type=int, default=0,
                    help="re-root the chain at the original DT after every N rounds "
                         "(0 = never); avoids getting trapped cycling among a few diagrams")
    ap.add_argument("--max-seconds", type=float, default=0.0,
                    help="stop generation after this many seconds (0 = no limit); resumable")
    ap.add_argument("--generate-only", action="store_true")
    ap.add_argument("--xlsx", default=None)
    ap.add_argument("--svg", default=None)
    ap.add_argument("--json", default=None)
    ap.add_argument("--max-draw", type=int, default=60)
    ap.add_argument("--check", action="append", default=[],
                    help="a DT code to test for membership among the sampled diagrams "
                         "(repeatable)")
    ap.add_argument("--check-file", default=None,
                    help="file with one DT code per line to test for membership")
    ap.add_argument("--verify", type=int, default=0, metavar="N",
                    help="confidence check: exact VF2 of each class rep against up to N "
                         "other distinct members (0 = off; use a large N or -1 for all)")
    ap.add_argument("--gui", action="store_true",
                    help="launch the graphical interface (also the default when no "
                         "arguments are given)")
    args = ap.parse_args(raw)

    if (not raw) or args.gui:
        launch_gui(args)
        return
    run_pipeline(args)


def run_pipeline(args, log=None):
    """Run generation -> dedup -> (verify) -> membership check -> score -> outputs.
    Prints progress with print(); the GUI redirects stdout to capture it."""
    t_start = time.time()
    print("Generating chain: %d rounds, backtrack %dx%d ..."
          % (args.rounds, args.backtrack_rounds, args.backtrack_steps), flush=True)
    chain = generate_chain(
        args.dt, args.rounds, args.backtrack_rounds, args.backtrack_steps,
        args.seed, checkpoint=args.checkpoint,
        max_seconds=(args.max_seconds or None), reset_every=args.reset_every,
    )
    done = len(chain) - 1
    print("Chain has %d/%d rounds (%d DT codes)." % (done, args.rounds, len(chain)), flush=True)
    if args.generate_only or done < args.rounds:
        if done < args.rounds:
            print("Not finished; re-run to resume from checkpoint %s." % args.checkpoint)
        return

    print("Deduplicating %d DT codes ..." % len(chain), flush=True)
    classes = dedup(chain)
    print("  -> %d distinct representative diagrams." % len(classes), flush=True)

    if args.verify:
        n_sample = 0 if args.verify < 0 else args.verify
        rep = verify_classes(classes, sample=n_sample)
        bad = [r for r in rep if not r["all_isomorphic"]]
        print("Verification (exact VF2, sample=%s): %s"
              % (n_sample or "ALL", "all classes internally consistent"
                 if not bad else "MERGE ERROR in classes %s" % [r["rep_id"] for r in bad]),
              flush=True)
        for r in rep:
            print("  class #%d: %d checked, all isomorphic=%s"
                  % (r["rep_id"], r["checked"], r["all_isomorphic"]), flush=True)

    queries = list(args.check)
    if args.check_file and os.path.exists(args.check_file):
        with open(args.check_file) as fh:
            queries += [ln.strip() for ln in fh if ln.strip()]
    check_results = check_sampled(classes, queries) if queries else []
    if check_results:
        print("Membership check (%d queries):" % len(check_results), flush=True)
        for r in check_results:
            if r.get("error"):
                print("  [error] %s : %s" % (r["dt"], r["error"]))
            elif r["sampled"]:
                print("  SAMPLED   -> rep #%d (mult %d, first seen round %d) : %s"
                      % (r["matches_rep_id"], r["multiplicity"], r["first_round"], r["dt"]))
            else:
                print("  NOT found -> %s" % r["dt"])

    print("Scoring representatives ...", flush=True)
    scored = score_representatives(classes)

    fps = set(m["linking_fp"] for m in scored if m["linking_fp"] is not None)
    print("  same-link check: %d distinct linking-number fingerprint(s) among "
          "representatives (expected 1: all are the same link by construction)"
          % len(fps), flush=True)

    run_info = {
        "software": "score_diagramV2_0.py",
        "root_DT": args.dt,
        "rounds": args.rounds,
        "backtrack_rounds": args.backtrack_rounds,
        "backtrack_steps": args.backtrack_steps,
        "reset_every": args.reset_every,
        "seed": args.seed,
        "total_DT_codes": len(chain),
        "distinct_representatives": len(classes),
        "distinct_linking_fingerprints": len(fps),
        "jones_polynomial": "populated only under Sage (SnapPy standalone cannot compute it)",
        "scored_on": "canonical DT of each class (labeling-independent, reproducible)",
        "dedup_equivalence": "signed diagram isomorphism (rotation/reflection/relabel/"
                             "cyclic-permutation/component-reorder; over-under preserved). "
                             "Fast signature = WL hash + strand-length + face-degree spectra; "
                             "exact backstops: --verify (VF2) and canonical_dt().",
        "scoring_layout_2d": "Tutte (unit-circle boundary)",
        "figure_layout_2d": "shaped-tutte, ellipse, aspect 1.0 (draw_dt panel and skeleton share it)",
        "layout_3d": "spherical Kamada-Kawai crossing centers (unit sphere)",
        "generated_utc": time.strftime("%Y-%m-%d %H:%M:%S UTC", time.gmtime()),
        "runtime_seconds": round(time.time() - t_start, 1),
    }

    if args.json:
        payload = {
            "run_info": run_info,
            "membership_check": check_results,
            "representatives": [
                {
                    "rep_id": m["_class"]["rep_id"],
                    "rank": m["rank"],
                    "dt": m["dt"],
                    "multiplicity": m["_class"]["multiplicity"],
                    "first_round": m["_class"]["rep_round"],
                    "jones": m["jones"],
                    "linking_fingerprint": list(m["linking_fp"]) if m["linking_fp"] else None,
                    "composite": m["composite"],
                    "quality": m["quality"],
                    "combinatorial": _strip_private({"c": m["combinatorial"]})["c"],
                    "graph": {k: v for k, v in m["graph"].items() if not k.startswith("_")},
                    "geom2d": {k: v for k, v in m["geom2d"].items() if not k.startswith("_")},
                    "sphere3d": {k: v for k, v in m["sphere3d"].items() if not k.startswith("_")},
                }
                for m in scored
            ],
        }
        with open(args.json, "w") as fh:
            json.dump(payload, fh, indent=2)
        print("wrote %s" % args.json)

    if args.xlsx:
        write_excel(scored, args.xlsx, run_info)
        print("wrote %s" % args.xlsx)
    if args.svg:
        make_figure(scored, args.svg, max_draw=args.max_draw)
        print("wrote %s" % args.svg)

    print("\nTop 3 representatives by composite:")
    for m in scored[:3]:
        print("  #%d  composite %.3f  strands %s  |Aut| %d  mult %d"
              % (m["_class"]["rep_id"], m["composite"],
                 m["combinatorial"]["strand_visit_lengths"],
                 m["graph"]["automorphism_order"], m["_class"]["multiplicity"]))


if __name__ == "__main__":
    main()
